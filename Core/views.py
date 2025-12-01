import json
from django.shortcuts import get_object_or_404, render, redirect
from django.db import models, transaction
from django.db.models import Avg, Sum, Count, Q, F
from django.db.models.functions import TruncDate
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash, get_user_model
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.core.paginator import Paginator
from django.core.serializers.json import DjangoJSONEncoder
from .models import Inventario, Material, Notificacion, Solicitud, DetalleSolicitud, Movimiento, Usuario, Alerta, MLResult, Local
from .forms import MaterialForm, MaterialInventarioForm, SolicitudForm, FiltroSolicitudesForm, CambiarPasswordForm, DetalleSolicitudFormSet, EditarMaterialForm, LocalForm
from .decorators import verificar_rol
from django.http import JsonResponse
from datetime import timedelta
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

User = get_user_model()

#----------------------------------------------------------------------------------------
# Vistas según roles
@login_required
def tecnico(request):
    return render(request, 'rol/tecnico.html')

@login_required
def historial_tecnico(request):
    """Vista del historial para técnicos - muestra sus solicitudes y movimientos"""
    
    # Obtener solicitudes del técnico
    solicitudes = Solicitud.objects.filter(
        solicitante=request.user
    ).prefetch_related('detalles__material').order_by('-fecha_solicitud')[:20]
    
    context = {
        'solicitudes': solicitudes,
    }
    
    return render(request, 'funcionalidad/historial_tecnico.html', context)

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def bodega(request):
    return render(request, 'rol/bodega.html')

@login_required
def chofer(request):
    return render(request, 'rol/chofer.html')

@login_required
@verificar_rol('GERENCIA')
def gerente(request):
    return render(request, 'rol/gerente.html')

#----------------------------------------------------------------------------------------
# ===================================== AUTENTICACIÓN ========================================
def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            next_url = request.GET.get('next', 'dashboard')
            return redirect(next_url)
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    
    return render(request, 'general/login.html')

def logout_view(request):
    storage = messages.get_messages(request)
    storage.used = True
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente')
    return redirect('login')

@login_required
def cambiar_password(request):
    if request.method == 'POST':
        form = CambiarPasswordForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, '¡Tu contraseña ha sido cambiada exitosamente!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = CambiarPasswordForm(user=request.user)
    
    return render(request, 'funcionalidad/cambiar_password.html', {'form': form})

#-------------------------------------------------------------------------------------------------------
# ======================================== INVENTARIO ========================================
@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])  # Solo BODEGA y GERENCIA
def inventario(request):
    query = request.GET.get('q', '').strip()
    items_por_pagina = request.GET.get('items', '10')
    
    try:
        items_por_pagina = int(items_por_pagina)
        if items_por_pagina not in [10, 25, 50, 100]:
            items_por_pagina = 25
    except ValueError:
        items_por_pagina = 25
    
    inventario_lista = Inventario.objects.select_related('material').all()
    
    if query:
        inventario_lista = inventario_lista.filter(
            Q(material__codigo__icontains=query) |
            Q(material__descripcion__icontains=query) |
            Q(material__ubicacion__icontains=query)
        )
    
    inventario_lista = inventario_lista.order_by('material__codigo')
    
    paginator = Paginator(inventario_lista, items_por_pagina)
    page_number = request.GET.get('page', 1)
    inventario = paginator.get_page(page_number)
    
    for item in inventario:
        promedio_ml = MLResult.objects.filter(
            material=item.material
        ).aggregate(
            promedio=Avg('stock_min_calculado')
        )['promedio']
        item.stock_critico_promedio = round(promedio_ml) if promedio_ml else None
    
    context = {
        'inventario': inventario,
        'total_items': paginator.count,
        'query': query,
        'items_por_pagina': items_por_pagina,
    }
    
    return render(request, 'funcionalidad/inv_inventario.html', context)

@login_required
def detalle_material(request, id):
    # TODOS pueden ver detalles si tienen acceso al inventario
    if request.user.rol not in ['BODEGA', 'GERENCIA']:
        messages.error(request, 'No tienes permiso para ver el inventario.')
        return redirect('dashboard')
    material = get_object_or_404(Material, id=id)
    return render(request, 'funcionalidad/inv_detalle_material.html', {'material': material})

@login_required
@verificar_rol('BODEGA')  # SOLO BODEGA puede editar
def editar_material(request, id):
    material = get_object_or_404(Material, id=id)
    
    if request.method == 'POST':
        form = EditarMaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, f'Material {material.codigo} actualizado exitosamente.')
            return redirect('inventario')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = EditarMaterialForm(instance=material)
    
    return render(request, 'funcionalidad/inv_editar_material.html', {
        'form': form,
        'material': material
    })

@login_required
@verificar_rol('BODEGA')  # SOLO BODEGA puede ingresar
def ingreso_material(request):
    last_material = Material.objects.order_by('-id').first()
    next_id = (last_material.id + 1) if last_material else 1
    codigo_sugerido = f"MAT{next_id:04d}"
    
    if request.method == 'POST':
        form = MaterialInventarioForm(request.POST)
        if form.is_valid():
            material = form.save(commit=False)
            if not material.codigo:
                material.codigo = codigo_sugerido
            material.save()
            
            Inventario.objects.create(
                material=material,
                stock_actual=form.cleaned_data['stock_actual'],
                stock_seguridad=form.cleaned_data['stock_seguridad'],
            )
            
            return redirect('inventario')
    else:
        form = MaterialInventarioForm(initial={'codigo': codigo_sugerido})
    
    return render(request, 'funcionalidad/inv_ingreso_material.html', {'form': form})

# ============================================= SOLICITUDES DE MATERIALES ====================================
@login_required
@verificar_rol('TECNICO')  # SOLO TÉCNICO ve sus solicitudes
def crear_solicitud(request):
    """Vista para crear una solicitud con múltiples materiales"""
    if request.method == 'POST':
        form = SolicitudForm(request.POST)
        formset = DetalleSolicitudFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            try:
                with transaction.atomic():
                    solicitud = form.save(commit=False)
                    solicitud.solicitante = request.user
                    solicitud.save()
                    
                    formset.instance = solicitud
                    formset.save()
                    
                    messages.success(
                        request,
                        f'Solicitud #{solicitud.id} creada exitosamente con {solicitud.total_items()} materiales.'
                    )
                    return redirect('mis_solicitudes')
            except Exception as e:
                messages.error(request, f'Error al crear la solicitud: {str(e)}')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = SolicitudForm()
        formset = DetalleSolicitudFormSet()
    
    context = {
        'form': form,
        'formset': formset,
    }
    
    return render(request, 'funcionalidad/solmat_crear_solicitud.html', context)

@login_required
def mis_solicitudes(request):
    """Vista para listar las solicitudes del usuario actual"""
    solicitudes = Solicitud.objects.filter(
        solicitante=request.user
    ).prefetch_related('detalles__material').order_by('-fecha_solicitud')
    
    context = {
        'solicitudes': solicitudes,
    }
    
    return render(request, 'funcionalidad/solmat_mis_solicitudes.html', context)

@login_required
def detalle_solicitud(request, solicitud_id):
    """Vista para ver el detalle completo de una solicitud"""
    solicitud = get_object_or_404(
        Solicitud.objects.prefetch_related('detalles__material'),
        id=solicitud_id
    )
    
    # Verificar permisos según rol
    if request.user.rol == 'TECNICO':
        # TÉCNICO solo ve las suyas
        if solicitud.solicitante != request.user:
            messages.error(request, 'No tienes permiso para ver esta solicitud.')
            return redirect('mis_solicitudes')
    elif request.user.rol in ['BODEGA', 'GERENCIA']:
        # BODEGA y GERENCIA ven todas
        pass
    else:
        messages.error(request, 'No tienes permiso para ver solicitudes.')
        return redirect('dashboard')
    
    context = {
        'solicitud': solicitud,
    }
    
    return render(request, 'funcionalidad/solmat_detalle_solicitud.html', context)

@login_required
@verificar_rol(['BODEGA']) # SOLO BODEGA puede gestionar
def gestionar_solicitudes(request):
    """Vista para que BODEGA gestione todas las solicitudes"""
    solicitudes_pendientes = Solicitud.objects.filter(
        estado='pendiente'
    ).prefetch_related('detalles__material').order_by('-fecha_solicitud')
    
    todas_solicitudes = Solicitud.objects.all().prefetch_related(
        'detalles__material'
    ).order_by('-fecha_solicitud')[:20]
    
    context = {
        'solicitudes_pendientes': solicitudes_pendientes,
        'todas_solicitudes': todas_solicitudes,
    }
    
    return render(request, 'funcionalidad/solmat_gestionar.html', context)

@login_required
@verificar_rol('BODEGA')
def aprobar_solicitud(request, solicitud_id):
    """Aprobar una solicitud"""
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)
    
    if request.method == 'POST':
        solicitud.estado = 'aprobada'
        solicitud.respondido_por = request.user
        solicitud.fecha_respuesta = timezone.now()
        solicitud.save()
        
        # CREAR NOTIFICACIÓN PARA EL TÉCNICO
        Notificacion.objects.create(
            usuario=solicitud.solicitante,
            tipo='solicitud_aprobada',
            mensaje=f'Tu solicitud #{solicitud.id} ha sido aprobada y está en proceso de despacho.',
            url=f'/solicitud/{solicitud.id}/'
        )
        
        messages.success(request, f'✓ Solicitud #{solicitud.id} aprobada exitosamente. Técnico notificado.')
        return redirect('gestionar_solicitudes')
    
    return redirect('detalle_solicitud', solicitud_id=solicitud_id)


@login_required
@verificar_rol('BODEGA')
def rechazar_solicitud(request, solicitud_id):
    """Rechazar una solicitud"""
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)
    
    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '')
        solicitud.estado = 'rechazada'
        solicitud.respondido_por = request.user
        solicitud.fecha_respuesta = timezone.now()
        solicitud.observaciones = observaciones
        solicitud.save()
        
        # CREAR NOTIFICACIÓN PARA EL TÉCNICO
        mensaje_notif = f'Tu solicitud #{solicitud.id} ha sido rechazada.'
        if observaciones:
            mensaje_notif += f' Motivo: {observaciones[:100]}'
        
        Notificacion.objects.create(
            usuario=solicitud.solicitante,
            tipo='solicitud_rechazada',
            mensaje=mensaje_notif,
            url=f'/solicitud/{solicitud.id}/'
        )
        
        messages.warning(request, f'⚠ Solicitud #{solicitud.id} rechazada. Técnico notificado.')
        return redirect('gestionar_solicitudes')
    
    return redirect('detalle_solicitud', solicitud_id=solicitud_id)


@login_required
@verificar_rol('TECNICO')  # SOLO el TÉCNICO puede cancelar las suyas
def cancelar_solicitud(request, solicitud_id):
    """Cancelar una solicitud propia (solo si está pendiente)"""
    solicitud = get_object_or_404(Solicitud, id=solicitud_id, solicitante=request.user)
    
    if solicitud.estado == 'pendiente' and request.method == 'POST':
        solicitud.delete()
        messages.info(request, 'Solicitud cancelada exitosamente.')
        return redirect('mis_solicitudes')
    
    messages.error(request, 'No se puede cancelar esta solicitud.')
    return redirect('detalle_solicitud', solicitud_id=solicitud_id)

@login_required
@verificar_rol(['BODEGA'])
def despachar_solicitud(request, solicitud_id):
    """Marcar una solicitud aprobada como despachada y descontar stock"""
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)
    
    if solicitud.estado != 'aprobada':
        messages.error(request, 'Solo se pueden despachar solicitudes aprobadas.')
        return redirect('gestionar_solicitudes')
    
    if request.method == 'POST':
        try:
            with transaction.atomic():
                for detalle in solicitud.detalles.all():
                    try:
                        inventario = detalle.material.inventario
                        
                        # Verificar stock suficiente
                        if inventario.stock_actual < detalle.cantidad:
                            messages.error(
                                request,
                                f'Stock insuficiente para {detalle.material.descripcion}. '
                                f'Disponible: {inventario.stock_actual}, Solicitado: {detalle.cantidad}'
                            )
                            return redirect('detalle_solicitud', solicitud_id=solicitud_id)
                        
                        # Descontar stock
                        stock_anterior = inventario.stock_actual
                        inventario.stock_actual -= detalle.cantidad
                        inventario.save()
                        
                        # Obtener usuario del modelo Usuario
                        try:
                            usuario_movimiento = Usuario.objects.get(email=request.user.email)
                        except Usuario.DoesNotExist:
                            usuario_movimiento = None
                        
                        # REGISTRAR UN SOLO MOVIMIENTO (SIN DUPLICAR)
                        if usuario_movimiento:
                            Movimiento.objects.create(
                                material=detalle.material,
                                usuario=usuario_movimiento,
                                solicitud=solicitud,
                                tipo='salida',
                                cantidad=detalle.cantidad,
                                detalle=f'Despacho solicitud #{solicitud.id} - {solicitud.motivo}'
                            )
                        else:
                            # Fallback si no existe el usuario en el modelo Usuario
                            Movimiento.objects.create(
                                material=detalle.material,
                                tipo='salida',
                                cantidad=detalle.cantidad,
                                motivo=f'Despacho solicitud #{solicitud.id}',
                                usuario=request.user  # Si tu modelo acepta esto
                            )
                    
                    except Inventario.DoesNotExist:
                        messages.error(
                            request,
                            f'El material {detalle.material.codigo} no tiene inventario asociado.'
                        )
                        return redirect('detalle_solicitud', solicitud_id=solicitud_id)
                
                # Cambiar estado de la solicitud
                solicitud.estado = 'despachada'
                solicitud.save()
                
                # CREAR NOTIFICACIÓN PARA EL TÉCNICO
                Notificacion.objects.create(
                    usuario=solicitud.solicitante,
                    tipo='solicitud_despachada',
                    mensaje=f'Tu solicitud #{solicitud.id} ha sido despachada y está lista para retirar.',
                    url=f'/solicitud/{solicitud.id}/'
                )
                
                messages.success(
                    request,
                    f'✓ Solicitud #{solicitud.id} despachada exitosamente. Stock actualizado y técnico notificado.'
                )
                return redirect('gestionar_solicitudes')
        
        except Exception as e:
            messages.error(request, f'Error al despachar solicitud: {str(e)}')
            return redirect('detalle_solicitud', solicitud_id=solicitud_id)
    
    context = {
        'solicitud': solicitud,
    }
    
    return render(request, 'funcionalidad/solmat_confirmar_despacho.html', context)


@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def historial_solicitudes(request):
    """
    Vista para ver historial completo de todas las solicitudes.
    Solo accesible para BODEGA y GERENCIA.
    """
    from datetime import datetime
    
    # Filtros desde GET
    estado_filtro = request.GET.get('estado', '')
    solicitante_filtro = request.GET.get('solicitante', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Consulta base con agregación de cantidad total
    solicitudes = Solicitud.objects.select_related(
        'solicitante', 'respondido_por'
    ).prefetch_related('detalles__material').annotate(
        total_materiales=Sum('detalles__cantidad')  # ← AGREGAR ESTO
    ).order_by('-fecha_solicitud')
    
    # Aplicar filtros
    if estado_filtro:
        solicitudes = solicitudes.filter(estado=estado_filtro)
    
    if solicitante_filtro:
        solicitudes = solicitudes.filter(solicitante__username=solicitante_filtro)
    
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            solicitudes = solicitudes.filter(fecha_solicitud__gte=fecha_desde_dt)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            solicitudes = solicitudes.filter(fecha_solicitud__lte=fecha_hasta_dt)
        except ValueError:
            pass
    
    # Paginación
    paginator = Paginator(solicitudes, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas generales
    stats = {
        'total': Solicitud.objects.count(),
        'pendientes': Solicitud.objects.filter(estado='pendiente').count(),
        'aprobadas': Solicitud.objects.filter(estado='aprobada').count(),
        'despachadas': Solicitud.objects.filter(estado='despachada').count(),
        'rechazadas': Solicitud.objects.filter(estado='rechazada').count(),
        'canceladas': Solicitud.objects.filter(estado='cancelada').count(),
    }
    
    # Lista de técnicos para el filtro
    tecnicos = User.objects.filter(rol='TECNICO').order_by('first_name')
    
    context = {
        'solicitudes': page_obj,
        'page_obj': page_obj,
        'stats': stats,
        'tecnicos': tecnicos,
        'estado_filtro': estado_filtro,
        'solicitante_filtro': solicitante_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'funcionalidad/solmat_historial.html', context)


# ============================================= Registro de movimientos ====================================
@login_required
@verificar_rol(['BODEGA']) # SOLO BODEGA puede registrar entradas
def registrar_entrada(request, material_id):
    """Registrar entrada manual de material"""
    material = get_object_or_404(Material, id=material_id)
    
    try:
        inventario = material.inventario
    except Inventario.DoesNotExist:
        messages.error(request, 'Este material no tiene inventario asociado.')
        return redirect('inventario')
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0))
        detalle = request.POST.get('detalle', '')
        
        if cantidad <= 0:
            messages.error(request, 'La cantidad debe ser mayor a 0.')
        else:
            try:
                with transaction.atomic():
                    stock_anterior = inventario.stock_actual
                    stock_nuevo = stock_anterior + cantidad
                    
                    inventario.stock_actual = stock_nuevo
                    inventario.save()
                    
                    try:
                        usuario_movimiento = Usuario.objects.get(
                            email=request.user.email
                        )
                    except Usuario.DoesNotExist:
                        messages.error(request, 'Tu usuario no está registrado en el sistema de personal.')
                        return redirect('detalle_material', id=material_id)
                    
                    Movimiento.objects.create(
                        material=material,
                        usuario=usuario_movimiento,
                        tipo='entrada',
                        cantidad=cantidad,
                        detalle=detalle or f'Entrada manual registrada por {request.user.username}'
                    )
                    
                    messages.success(
                        request,
                        f'Entrada registrada: +{cantidad} {material.unidad_medida}. Stock actual: {stock_nuevo}'
                    )
                    return redirect('detalle_material', id=material_id)
            except Exception as e:
                messages.error(request, f'Error al registrar entrada: {str(e)}')
    
    context = {
        'material': material,
        'inventario': inventario,
    }
    
    return render(request, 'funcionalidad/mov_registrar_entrada.html', context)

@login_required
@verificar_rol('BODEGA')  # SOLO BODEGA puede registrar salidas
def registrar_salida(request, material_id):
    """Registrar salida manual de material"""
    material = get_object_or_404(Material, id=material_id)
    
    try:
        inventario = material.inventario
    except Inventario.DoesNotExist:
        messages.error(request, 'Este material no tiene inventario asociado.')
        return redirect('inventario')
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0))
        detalle = request.POST.get('detalle', '')
        
        if cantidad <= 0:
            messages.error(request, 'La cantidad debe ser mayor a 0.')
        elif cantidad > inventario.stock_actual:
            messages.error(request, f'Stock insuficiente. Disponible: {inventario.stock_actual}')
        else:
            try:
                with transaction.atomic():
                    stock_anterior = inventario.stock_actual
                    stock_nuevo = stock_anterior - cantidad
                    
                    inventario.stock_actual = stock_nuevo
                    inventario.save()
                    
                    try:
                        usuario_movimiento = Usuario.objects.get(
                            email=request.user.email
                        )
                    except Usuario.DoesNotExist:
                        messages.error(request, 'Tu usuario no está registrado en el sistema de personal.')
                        return redirect('detalle_material', id=material_id)
                    
                    Movimiento.objects.create(
                        material=material,
                        usuario=usuario_movimiento,
                        tipo='salida',
                        cantidad=cantidad,
                        detalle=detalle or f'Salida manual registrada por {request.user.username}'
                    )
                    
                    messages.success(
                        request,
                        f'Salida registrada: -{cantidad} {material.unidad_medida}. Stock actual: {stock_nuevo}'
                    )
                    return redirect('detalle_material', id=material_id)
            except Exception as e:
                messages.error(request, f'Error al registrar salida: {str(e)}')
    
    context = {
        'material': material,
        'inventario': inventario,
    }
    
    return render(request, 'funcionalidad/mov_registrar_salida.html', context)

@login_required
@verificar_rol('BODEGA')  # SOLO BODEGA puede ajustar
def ajustar_inventario(request, material_id):
    """Ajustar inventario (correcciones)"""
    material = get_object_or_404(Material, id=material_id)
    
    try:
        inventario = material.inventario
    except Inventario.DoesNotExist:
        messages.error(request, 'Este material no tiene inventario asociado.')
        return redirect('inventario')
    
    if request.method == 'POST':
        nuevo_stock = int(request.POST.get('nuevo_stock', 0))
        detalle = request.POST.get('detalle', '')
        
        if nuevo_stock < 0:
            messages.error(request, 'El stock no puede ser negativo.')
        else:
            try:
                with transaction.atomic():
                    stock_anterior = inventario.stock_actual
                    diferencia = nuevo_stock - stock_anterior
                    
                    inventario.stock_actual = nuevo_stock
                    inventario.save()
                    
                    try:
                        usuario_movimiento = Usuario.objects.get(
                            email=request.user.email
                        )
                    except Usuario.DoesNotExist:
                        messages.error(request, 'Tu usuario no está registrado en el sistema de personal.')
                        return redirect('detalle_material', id=material_id)
                    
                    Movimiento.objects.create(
                        material=material,
                        usuario=usuario_movimiento,
                        tipo='ajuste',
                        cantidad=abs(diferencia),
                        detalle=f'Ajuste de inventario ({diferencia:+d}): {detalle}'
                    )
                    
                    messages.success(
                        request,
                        f'Inventario ajustado. Stock anterior: {stock_anterior}, Stock nuevo: {nuevo_stock}'
                    )
                    return redirect('detalle_material', id=material_id)
            except Exception as e:
                messages.error(request, f'Error al ajustar inventario: {str(e)}')
    
    context = {
        'material': material,
        'inventario': inventario,
    }
    
    return render(request, 'funcionalidad/mov_ajustar_inventario.html', context)

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])  # Ambos pueden ver historial
def historial_movimientos(request, material_id):
    """Ver historial de movimientos de un material"""
    material = get_object_or_404(Material, id=material_id)
    movimientos = Movimiento.objects.filter(
        material=material
    ).select_related('usuario').order_by('-fecha')
    
    context = {
        'material': material,
        'movimientos': movimientos,
    }
    
    return render(request, 'funcionalidad/mov_historial.html', context)

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])  # Ambos pueden ver historial global
def historial_movimientos_global(request):
    """Historial completo de todos los movimientos del sistema"""
    movimientos = Movimiento.objects.select_related(
        'material', 'usuario'
    ).all().order_by('-fecha')[:100]
    
    context = {
        'movimientos': movimientos
    }
    
    return render(request, 'funcionalidad/mov_historial_global.html', context)


@login_required
def solicitud(request):
    """
    Vista simple de solicitud.
    Redirige a crear solicitud.
    """
    return redirect('crear_solicitud')

# ============================================= NOTIFICACIONES ====================================

@login_required
def mis_notificaciones(request):
    """
    Vista para listar todas las notificaciones del usuario.
    
    - TÉCNICO: Solo notificaciones de sus solicitudes
    - BODEGA/GERENCIA: Todas las notificaciones
    """
    usuario = request.user
    rol = usuario.rol
    
    # Filtrar notificaciones según el rol
    if rol == 'TECNICO':
        notificaciones = Notificacion.objects.filter(
            usuario=usuario,
            tipo__in=['solicitud_aprobada', 'solicitud_rechazada', 'solicitud_despachada']
        ).order_by('-creada_en')
    elif rol in ['BODEGA', 'GERENCIA']:
        notificaciones = Notificacion.objects.filter(
            usuario=usuario
        ).order_by('-creada_en')
    else:
        notificaciones = Notificacion.objects.none()
    
    # Paginación
    paginator = Paginator(notificaciones, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'total': notificaciones.count(),
        'no_leidas': notificaciones.filter(leida=False).count(),
        'leidas': notificaciones.filter(leida=True).count(),
    }
    
    context = {
        'notificaciones': page_obj,
        'page_obj': page_obj,
        'stats': stats,
    }
    
    return render(request, 'funcionalidad/notificaciones.html', context)

@login_required
def marcar_leida(request, id):
    """Marcar una notificación como leída"""
    notificacion = get_object_or_404(Notificacion, id=id, usuario=request.user)
    notificacion.leida = True
    notificacion.save()
    
    # Si tiene URL asociada, redirigir ahí
    if notificacion.url:
        return redirect(notificacion.url)
    
    return redirect('mis_notificaciones')

@login_required
def marcar_todas_leidas(request):
    """Marcar todas las notificaciones del usuario como leídas"""
    if request.method == 'POST':
        count = Notificacion.objects.filter(
            usuario=request.user,
            leida=False
        ).update(leida=True)
        
        messages.success(request, f'✓ {count} notificaciones marcadas como leídas.')
    
    return redirect('mis_notificaciones')

@login_required
def eliminar_notificacion(request, id):
    """Eliminar una notificación"""
    notificacion = get_object_or_404(Notificacion, id=id, usuario=request.user)
    
    if request.method == 'POST':
        notificacion.delete()
        messages.success(request, '✓ Notificación eliminada.')
    
    return redirect('mis_notificaciones')

@login_required
def obtener_notificaciones_json(request):
    """
    API endpoint para obtener notificaciones no leídas en formato JSON.
    Usado por el dropdown del navbar.
    
    - TÉCNICO: Solo notificaciones de sus solicitudes
    - BODEGA/GERENCIA: Todas las notificaciones del sistema
    """
    usuario = request.user
    rol = usuario.rol
    
    # Filtrar notificaciones según el rol
    if rol == 'TECNICO':
        # TÉCNICO: Solo notificaciones relacionadas a SUS solicitudes
        notificaciones = Notificacion.objects.filter(
            usuario=usuario,
            leida=False,
            tipo__in=['solicitud_aprobada', 'solicitud_rechazada', 'solicitud_despachada']
        ).order_by('-creada_en')[:10]
    elif rol in ['BODEGA', 'GERENCIA']:
        # BODEGA y GERENCIA: Todas las notificaciones
        notificaciones = Notificacion.objects.filter(
            usuario=usuario,
            leida=False
        ).order_by('-creada_en')[:10]
    else:
        # Otro rol: Sin notificaciones
        notificaciones = Notificacion.objects.none()
    
    # Contar total de no leídas
    if rol == 'TECNICO':
        count_total = Notificacion.objects.filter(
            usuario=usuario,
            leida=False,
            tipo__in=['solicitud_aprobada', 'solicitud_rechazada', 'solicitud_despachada']
        ).count()
    elif rol in ['BODEGA', 'GERENCIA']:
        count_total = Notificacion.objects.filter(
            usuario=usuario,
            leida=False
        ).count()
    else:
        count_total = 0
    
    data = {
        'count': count_total,
        'notificaciones': [
            {
                'id': n.id,
                'tipo': n.tipo,
                'mensaje': n.mensaje,
                'url': n.url or '',
                'fecha': n.creada_en.strftime('%d/%m/%Y %H:%M'),
                'icono': get_icono_notificacion(n.tipo)
            }
            for n in notificaciones
        ]
    }
    
    return JsonResponse(data)

def get_icono_notificacion(tipo):
    """Retorna el ícono FontAwesome según el tipo de notificación"""
    iconos = {
        'stock_critico': 'fa-exclamation-triangle',
        'solicitud_pendiente': 'fa-clipboard-list',
        'material_nuevo': 'fa-box',
        'aprobacion': 'fa-check-circle',
        'solicitud_aprobada': 'fa-check-circle',      # Para técnicos
        'solicitud_rechazada': 'fa-times-circle',     # Para técnicos
        'solicitud_despachada': 'fa-truck',           # Para técnicos
    }
    return iconos.get(tipo, 'fa-bell')


# ============================================= DASHBOARD ====================================
@login_required
def dashboard(request):
    """Dashboard principal con estadísticas adaptadas según el rol del usuario."""
    usuario = request.user
    rol = usuario.rol
    
    # KPIs BASE para todos los roles
    total_materiales = Material.objects.count()
    solicitudes_totales = Solicitud.objects.count()
    solicitudes_pendientes = Solicitud.objects.filter(estado='pendiente').count()
    
    # Materiales críticos
    materiales_criticos_lista = Inventario.objects.filter(
        stock_actual__lte=F('stock_seguridad')
    ).select_related('material')[:5]
    materiales_criticos = Inventario.objects.filter(
        stock_actual__lte=F('stock_seguridad')
    ).count()
    
    # Context base
    context = {
        'usuario': usuario,
        'rol': rol,
        'total_materiales': total_materiales,
        'solicitudes_totales': solicitudes_totales,
        'solicitudes_pendientes': solicitudes_pendientes,
        'materiales_criticos': materiales_criticos,
        'materiales_criticos_lista': materiales_criticos_lista,
    }
    
    # ========== TÉCNICO ==========
    if rol == 'TECNICO':
        mis_solicitudes = Solicitud.objects.filter(solicitante=usuario)
        
        # Stats para gráfico de solicitudes
        solicitudes_data = [
            mis_solicitudes.filter(estado='pendiente').count(),
            mis_solicitudes.filter(estado='aprobada').count(),
            mis_solicitudes.filter(estado='rechazada').count(),
            mis_solicitudes.filter(estado='despachada').count(),
        ]
        
        context.update({
            'mis_solicitudes_count': mis_solicitudes.count(),
            'mis_pendientes': mis_solicitudes.filter(estado='pendiente').count(),
            'mis_aprobadas': mis_solicitudes.filter(estado='aprobada').count(),
            'solicitudes_recientes': mis_solicitudes.order_by('-fecha_solicitud')[:5],
            
            # Top materiales del técnico
            'top_materiales': DetalleSolicitud.objects.filter(
                solicitud__solicitante=usuario
            ).values(
                'material__codigo', 'material__descripcion'
            ).annotate(
                total_solicitado=Sum('cantidad')
            ).order_by('-total_solicitado')[:5],
            
            # Stats para gráfico - JSON serializado
            'solicitudes_labels': json.dumps(['Pendientes', 'Aprobadas', 'Rechazadas', 'Despachadas']),
            'solicitudes_data': json.dumps(solicitudes_data),
        })
    
    # ========== BODEGA ==========
    elif rol == 'BODEGA':
        # Stock total
        stock_total = Inventario.objects.aggregate(
            total=Sum('stock_actual')
        )['total'] or 0
        
        # Stats para gráfico de solicitudes
        solicitudes_data = [
            Solicitud.objects.filter(estado='pendiente').count(),
            Solicitud.objects.filter(estado='aprobada').count(),
            Solicitud.objects.filter(estado='rechazada').count(),
            Solicitud.objects.filter(estado='despachada').count(),
        ]
        
        # ========== PROCESAR MOVIMIENTOS DE LOS ÚLTIMOS 7 DÍAS ==========
        movimientos_7dias = Movimiento.objects.filter(
            fecha__gte=timezone.now() - timedelta(days=7)
        ).values('fecha__date', 'tipo').annotate(
            total=Sum('cantidad')
        ).order_by('fecha__date')
        
        # Agrupar por fecha y tipo
        mov_dict = defaultdict(lambda: {'entrada': 0, 'salida': 0})
        for mov in movimientos_7dias:
            fecha = mov['fecha__date'].strftime('%d/%m')
            mov_dict[fecha][mov['tipo']] = mov['total']
        
        # Obtener últimos 7 días
        fechas = []
        entradas = []
        salidas = []
        for i in range(6, -1, -1):
            fecha = (timezone.now() - timedelta(days=i)).date()
            fecha_str = fecha.strftime('%d/%m')
            fechas.append(fecha_str)
            entradas.append(mov_dict[fecha_str]['entrada'])
            salidas.append(mov_dict[fecha_str]['salida'])
        
        context.update({
            'stock_total': stock_total,
            'solicitudes_recientes': Solicitud.objects.select_related('solicitante').order_by('-fecha_solicitud')[:5],
            
            # Top materiales del sistema
            'top_materiales': DetalleSolicitud.objects.values(
                'material__codigo', 'material__descripcion'
            ).annotate(
                total_solicitado=Sum('cantidad')
            ).order_by('-total_solicitado')[:5],
            
            # Actividad reciente
            'actividad_reciente': Movimiento.objects.select_related(
                'material', 'usuario'
            ).order_by('-fecha')[:10],
            
            # Stats para gráficos - JSON serializado
            'solicitudes_labels': json.dumps(['Pendientes', 'Aprobadas', 'Rechazadas', 'Despachadas']),
            'solicitudes_data': json.dumps(solicitudes_data),
            
            # Datos para gráfico de movimientos
            'movimientos_fechas': json.dumps(fechas),
            'movimientos_entradas': json.dumps(entradas),
            'movimientos_salidas': json.dumps(salidas),
        })
    
    # ========== GERENCIA ==========
    elif rol == 'GERENCIA':
        total_usuarios = User.objects.filter(is_active=True).count()
        stock_total = Inventario.objects.aggregate(
            total=Sum('stock_actual')
        )['total'] or 0
        
        # Stats para gráfico de solicitudes
        solicitudes_data = [
            Solicitud.objects.filter(estado='pendiente').count(),
            Solicitud.objects.filter(estado='aprobada').count(),
            Solicitud.objects.filter(estado='rechazada').count(),
            Solicitud.objects.filter(estado='despachada').count(),
        ]
        
        # ========== PROCESAR MOVIMIENTOS DE LOS ÚLTIMOS 7 DÍAS ==========
        movimientos_7dias = Movimiento.objects.filter(
            fecha__gte=timezone.now() - timedelta(days=7)
        ).values('fecha__date', 'tipo').annotate(
            total=Sum('cantidad')
        ).order_by('fecha__date')
        
        # Agrupar por fecha y tipo
        mov_dict = defaultdict(lambda: {'entrada': 0, 'salida': 0})
        for mov in movimientos_7dias:
            fecha = mov['fecha__date'].strftime('%d/%m')
            mov_dict[fecha][mov['tipo']] = mov['total']
        
        # Obtener últimos 7 días
        fechas = []
        entradas = []
        salidas = []
        for i in range(6, -1, -1):
            fecha = (timezone.now() - timedelta(days=i)).date()
            fecha_str = fecha.strftime('%d/%m')
            fechas.append(fecha_str)
            entradas.append(mov_dict[fecha_str]['entrada'])
            salidas.append(mov_dict[fecha_str]['salida'])
        
        context.update({
            'total_usuarios': total_usuarios,
            'stock_total': stock_total,
            'solicitudes_recientes': Solicitud.objects.select_related('solicitante').order_by('-fecha_solicitud')[:5],
            
            # Top materiales del sistema
            'top_materiales': DetalleSolicitud.objects.values(
                'material__codigo', 'material__descripcion'
            ).annotate(
                total_solicitado=Sum('cantidad')
            ).order_by('-total_solicitado')[:5],
            
            # Actividad reciente
            'actividad_reciente': Movimiento.objects.select_related(
                'material', 'usuario'
            ).order_by('-fecha')[:10],
            
            # Stats para gráficos - JSON serializado
            'solicitudes_labels': json.dumps(['Pendientes', 'Aprobadas', 'Rechazadas', 'Despachadas']),
            'solicitudes_data': json.dumps(solicitudes_data),
            
            # Datos para gráfico de movimientos
            'movimientos_fechas': json.dumps(fechas),
            'movimientos_entradas': json.dumps(entradas),
            'movimientos_salidas': json.dumps(salidas),
        })
    
    # ========== MATERIALES POR CATEGORÍA (para todos) ==========
    materiales_por_categoria = Material.objects.values('categoria').annotate(
        total=Count('id')
    ).order_by('-total')
    
    categorias_labels = [dict(Material.CATEGORIA_CHOICES).get(item['categoria'], item['categoria']) 
                         for item in materiales_por_categoria]
    categorias_data = [item['total'] for item in materiales_por_categoria]
    
    context.update({
        'categorias_labels': json.dumps(categorias_labels),
        'categorias_data': json.dumps(categorias_data),
    })
    
    return render(request, 'general/dashboard.html', context)



# ============================================= EXPORTACIONES A EXCEL ====================================

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])  # Ambos pueden exportar
def exportar_inventario_excel(request):
    """Exportar inventario completo a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Headers
    headers = ['Código', 'Descripción', 'Categoría', 'Stock Actual', 'Stock Seguridad', 'Estado', 'Ubicación']
    ws.append(headers)
    
    # Estilo para headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    inventario = Inventario.objects.select_related('material').all()
    for inv in inventario:
        if inv.stock_actual <= inv.stock_seguridad:
            estado = 'CRÍTICO'
        else:
            estado = 'NORMAL'
        
        ws.append([
            inv.material.codigo,
            inv.material.descripcion,
            inv.material.get_categoria_display(),
            inv.stock_actual,
            inv.stock_seguridad,
            estado,
            inv.material.ubicacion or 'No especificada'
        ])
    
    # Ajustar anchos
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=inventario_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def exportar_solicitudes_excel(request):
    """Exportar solicitudes a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes"
    
    # Headers
    headers = ['ID', 'Fecha', 'Solicitante', 'Estado', 'Total Items', 'Motivo']
    ws.append(headers)
    
    # Estilo para headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    solicitudes = Solicitud.objects.select_related('solicitante').prefetch_related('detalles').order_by('-fecha_solicitud')[:100]
    for sol in solicitudes:
        ws.append([
            sol.id,
            sol.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
            sol.solicitante.username,
            sol.get_estado_display(),
            sol.total_items(),
            sol.motivo[:50]
        ])
    
    # Ajustar anchos
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=solicitudes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def exportar_movimientos_excel(request):
    """Exportar movimientos a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"
    
    # Headers
    headers = ['Fecha', 'Material', 'Tipo', 'Cantidad', 'Usuario', 'Detalle']
    ws.append(headers)
    
    # Estilo para headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    movimientos = Movimiento.objects.select_related('material', 'usuario').order_by('-fecha')[:200]
    for mov in movimientos:
        ws.append([
            mov.fecha.strftime('%d/%m/%Y %H:%M'),
            f"{mov.material.codigo} - {mov.material.descripcion}",
            mov.tipo.upper(),
            mov.cantidad,
            f"{mov.usuario.nombre} {mov.usuario.apellido}" if mov.usuario else 'Sistema',
            mov.detalle[:60]
        ])
    
    # Ajustar anchos
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=movimientos_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def exportar_reporte_completo_excel(request):
    """Reporte completo con múltiples hojas: Inventario, Solicitudes, Movimientos"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # ========== HOJA 1: INVENTARIO ==========
    ws_inv = wb.create_sheet("Inventario")
    headers_inv = ['Código', 'Descripción', 'Categoría', 'Stock Actual', 'Stock Seguridad', 'Estado']
    ws_inv.append(headers_inv)
    
    # Estilo headers
    for cell in ws_inv[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    inventario = Inventario.objects.select_related('material').all()
    for inv in inventario:
        estado = 'CRÍTICO' if inv.stock_actual <= inv.stock_seguridad else 'NORMAL'
        ws_inv.append([
            inv.material.codigo,
            inv.material.descripcion,
            inv.material.get_categoria_display(),
            inv.stock_actual,
            inv.stock_seguridad,
            estado
        ])
    
    # ========== HOJA 2: SOLICITUDES ==========
    ws_sol = wb.create_sheet("Solicitudes")
    headers_sol = ['ID', 'Fecha', 'Solicitante', 'Estado', 'Items']
    ws_sol.append(headers_sol)
    
    for cell in ws_sol[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    solicitudes = Solicitud.objects.select_related('solicitante').prefetch_related('detalles').order_by('-fecha_solicitud')[:100]
    for sol in solicitudes:
        ws_sol.append([
            sol.id,
            sol.fecha_solicitud.strftime('%d/%m/%Y'),
            sol.solicitante.username,
            sol.get_estado_display(),
            sol.total_items()
        ])
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_completo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


 # ==================================== GESTION DE LOCALES ==============================================

@login_required
@verificar_rol('GERENCIA')
def gestion_locales(request):
    """
    Vista para gestión de locales (solo GERENCIA).
    Lista todos los locales con búsqueda y paginación.
    """
    locales = Local.objects.all().order_by('codigo', 'nombre')
    
    # Búsqueda
    query = request.GET.get('q', '').strip()
    if query:
        locales = locales.filter(
            Q(nombre__icontains=query) | 
            Q(codigo__icontains=query) |
            Q(comuna__icontains=query)
        )
    
    # Paginación
    paginator = Paginator(locales, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'locales': page_obj,
        'page_obj': page_obj,
        'query': query,
        'total_locales': locales.count()
    }
    return render(request, 'funcionalidad/local_gestion.html', context)

@login_required
@verificar_rol('GERENCIA')
def local_crear(request):
    """
    Vista para crear un nuevo local (solo GERENCIA).
    """
    if request.method == 'POST':
        form = LocalForm(request.POST)
        if form.is_valid():
            local = form.save()
            messages.success(
                request, 
                f'✓ Local {local.codigo} - {local.nombre} creado exitosamente.'
            )
            return redirect('gestion_locales')
    else:
        form = LocalForm()
    
    context = {
        'form': form,
        'titulo': 'Agregar Local',
        'accion': 'Crear'
    }
    return render(request, 'funcionalidad/local_form.html', context)

@login_required
@verificar_rol('GERENCIA')
def local_editar(request, local_id):
    """
    Vista para editar un local existente (solo GERENCIA).
    """
    local = get_object_or_404(Local, pk=local_id)
    
    if request.method == 'POST':
        form = LocalForm(request.POST, instance=local)
        if form.is_valid():
            form.save()
            messages.success(
                request, 
                f'✓ Local {local.codigo} - {local.nombre} actualizado exitosamente.'
            )
            return redirect('gestion_locales')
    else:
        form = LocalForm(instance=local)
    
    context = {
        'form': form,
        'local': local,
        'titulo': 'Editar Local',
        'accion': 'Actualizar'
    }
    return render(request, 'funcionalidad/local_form.html', context)

@login_required
@verificar_rol('GERENCIA')
def local_eliminar(request, local_id):
    """
    Vista para eliminar un local (solo GERENCIA).
    Verifica que no tenga solicitudes asociadas.
    """
    local = get_object_or_404(Local, pk=local_id)
    
    if request.method == 'POST':
        # Verificar si tiene solicitudes asociadas
        if local.solicitudes.exists():
            messages.error(
                request,
                f'No se puede eliminar el local {local.codigo} - {local.nombre} '
                f'porque tiene {local.solicitudes.count()} solicitudes asociadas.'
            )
            return redirect('gestion_locales')
        
        nombre = str(local)
        local.delete()
        messages.success(request, f'✓ Local {nombre} eliminado exitosamente.')
        return redirect('gestion_locales')
    
    context = {
        'local': local,
        'solicitudes_count': local.solicitudes.count()
    }
    return render(request, 'funcionalidad/local_confirm_delete.html', context)

