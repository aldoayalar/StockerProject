import json
import pandas as pd
from django.shortcuts import get_object_or_404, render, redirect
from django.db import models, transaction
from django.db.models import Avg, Sum, Count, Q, F
from django.db.models.functions import TruncDate
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout, update_session_auth_hash, get_user_model
from django.contrib.auth.decorators import login_required
from django.utils import timezone
from django.urls import reverse
from django.core.paginator import Paginator
from django.core.serializers.json import DjangoJSONEncoder
from .models import Inventario, Material, Notificacion, Solicitud, DetalleSolicitud, Movimiento, Usuario, Alerta, MLResult, Local
from .forms import (MaterialForm, MaterialInventarioForm, SolicitudForm, FiltroSolicitudesForm, CambiarPasswordForm, 
                    DetalleSolicitudFormSet, EditarMaterialForm, LocalForm, CargaMasivaStockForm, UsuarioForm)
from .decorators import verificar_rol
from .services.ml_service import ejecutar_calculo_global
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from datetime import timedelta, datetime
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict




User = get_user_model()

#----------------------------------------------------------------------------------------
# Vistas según roles
@login_required
def tecnico(request):
    return render(request, 'rol/tecnico.html')

@login_required
def historial_tecnico(request):
    """Vista del historial para técnicos - muestra sus solicitudes y movimientos"""
    
    # Obtener solicitudes del técnico
    solicitudes = Solicitud.objects.filter(
        solicitante=request.user
    ).prefetch_related('detalles__material').order_by('-fecha_solicitud')[:20]
    
    context = {
        'solicitudes': solicitudes,
    }
    
    return render(request, 'funcionalidad/historial_tecnico.html', context)

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def bodega(request):
    return render(request, 'rol/bodega.html')

@login_required
def chofer(request):
    return render(request, 'rol/chofer.html')

@login_required
@verificar_rol('GERENCIA')
def gerente(request):
    return render(request, 'rol/gerente.html')

#----------------------------------------------------------------------------------------
# ===================================== AUTENTICACIÓN ========================================
def login_view(request):
    if request.user.is_authenticated:
        if request.user.is_superuser or getattr(request.user, 'rol', None) == 'SISTEMA':
            return redirect('sistema_home')   
        return redirect('dashboard')          

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            # Si es sistema/superuser -> siempre a su inicio
            if user.is_superuser or getattr(user, 'rol', None) == 'SISTEMA':
                return redirect('sistema_home')

            next_url = request.GET.get('next')
            if next_url:
                return redirect(next_url)
            return redirect('dashboard')

        # Credenciales inválidas (opcional)
        messages.error(request, '[translate:Usuario o contraseña incorrectos]')

    return render(request, 'general/login.html')

def logout_view(request):
    storage = messages.get_messages(request)
    storage.used = True
    logout(request)
    messages.success(request, 'Has cerrado sesión exitosamente')
    return redirect('login')

@login_required
def cambiar_password(request):
    if request.method == 'POST':
        form = CambiarPasswordForm(user=request.user, data=request.POST)
        if form.is_valid():
            user = form.save()
            update_session_auth_hash(request, user)
            messages.success(request, '¡Tu contraseña ha sido cambiada exitosamente!')
            return redirect('dashboard')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = CambiarPasswordForm(user=request.user)
    
    return render(request, 'funcionalidad/cambiar_password.html', {'form': form})

#-------------------------------------------------------------------------------------------------------
# ======================================== INVENTARIO ========================================
@login_required
@verificar_rol(['BODEGA', 'GERENCIA']) 
def inventario(request):
    query = request.GET.get('q', '').strip()
    items_por_pagina = request.GET.get('items', '10')
    
    try:
        items_por_pagina = int(items_por_pagina)
        if items_por_pagina not in [10, 25, 50, 100]:
            items_por_pagina = 25
    except ValueError:
        items_por_pagina = 25
    
    inventario_lista = Inventario.objects.select_related('material').all()
    
    if query:
        inventario_lista = inventario_lista.filter(
            Q(material__codigo__icontains=query) |
            Q(material__descripcion__icontains=query) |
            Q(material__ubicacion__icontains=query)
        )
    
    inventario_lista = inventario_lista.order_by('material__codigo')
    
    paginator = Paginator(inventario_lista, items_por_pagina)
    page_number = request.GET.get('page', 1)
    inventario = paginator.get_page(page_number)
    
    for item in inventario:
        promedio_ml = MLResult.objects.filter(
            material=item.material
        ).aggregate(
            promedio=Avg('stock_min_calculado')
        )['promedio']
        item.stock_critico_promedio = round(promedio_ml) if promedio_ml else None
    
    context = {
        'inventario': inventario,
        'total_items': paginator.count,
        'query': query,
        'items_por_pagina': items_por_pagina,
    }
    
    return render(request, 'funcionalidad/inv_inventario.html', context)

@login_required
def detalle_material(request, id):
    if request.user.rol not in ['BODEGA', 'GERENCIA']:
        messages.error(request, 'No tienes permiso para ver el inventario.')
        return redirect('inventario')

    material = get_object_or_404(Material, id=id)

    # Inventario asociado (puede no existir si algo quedó inconsistente)
    inventario = Inventario.objects.filter(material=material).first()

    # Últimos 10 movimientos de este material
    movimientos_recientes = Movimiento.objects.filter(
        material=material
    ).select_related('usuario').order_by('-fecha')[:10]

    context = {
        'material': material,
        'inventario': inventario,
        'movimientos_recientes': movimientos_recientes,
    }
    return render(request, 'funcionalidad/inv_detalle_material.html', context)

@login_required
@verificar_rol('BODEGA')  # SOLO BODEGA puede editar
def editar_material(request, id):
    material = get_object_or_404(Material, id=id)
    
    if request.method == 'POST':
        form = EditarMaterialForm(request.POST, instance=material)
        if form.is_valid():
            form.save()
            messages.success(request, f'Material {material.codigo} actualizado exitosamente.')
            return redirect('inventario')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = EditarMaterialForm(instance=material)
    
    return render(request, 'funcionalidad/inv_editar_material.html', {
        'form': form,
        'material': material
    })

@login_required
@verificar_rol('BODEGA')  # SOLO BODEGA puede ingresar
def ingreso_material(request):
    last_material = Material.objects.order_by('-id').first()
    next_id = (last_material.id + 1) if last_material else 1
    codigo_sugerido = f"MAT{next_id:04d}"
    
    if request.method == 'POST':
        form = MaterialInventarioForm(request.POST)
        if form.is_valid():
            material = form.save(commit=False)
            if not material.codigo:
                material.codigo = codigo_sugerido
            material.save()
            
            Inventario.objects.create(
                material=material,
                stock_actual=form.cleaned_data['stock_actual'],
                stock_seguridad=form.cleaned_data['stock_seguridad'],
            )
            
            return redirect('inventario')
    else:
        form = MaterialInventarioForm(initial={'codigo': codigo_sugerido})
    
    return render(request, 'funcionalidad/inv_ingreso_material.html', {'form': form})

@login_required
@verificar_rol(['BODEGA'])
def crear_material(request):
    if request.method == 'POST':
        form = MaterialForm(request.POST)
        if form.is_valid():
            material = form.save()
            # Crear inventario asociado en 0 si no existe
            Inventario.objects.get_or_create(
                material=material,
                defaults={
                    'stock_actual': 0,
                    'stock_seguridad': 0,
                }
            )
            messages.success(request, 'Material creado correctamente.')
            return redirect('inventario')  # tu vista de listado de inventario
    else:
        form = MaterialForm()

    return render(request, 'funcionalidad/material_form.html', {'form': form})

@login_required
@verificar_rol(['BODEGA'])
def carga_masiva_stock(request):
    if request.method == 'POST':
        form = CargaMasivaStockForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES['archivo']
            modo = form.cleaned_data['modo']

            try:
                df = pd.read_excel(archivo, engine='openpyxl')
            except Exception:
                messages.error(request, "No se pudo leer el archivo. Verifica que sea un Excel válido (.xlsx).")
                return redirect('carga_masiva_stock')

            # Debug rápido: cuántas filas y qué columnas vienen
            print("DF rows:", len(df), "cols:", list(df.columns))

            # Normalizar nombres de columnas a str
            df.columns = df.columns.map(str)

            if 'Código' not in df.columns or 'NuevoStock' not in df.columns:
                messages.error(request, "El archivo debe tener las columnas 'Código' y 'NuevoStock'.")
                return redirect('carga_masiva_stock')

            actualizados = 0
            no_encontrados = []

            with transaction.atomic():
                for _, row in df.iterrows():
                    codigo = str(row['Código']).strip()
                    if not codigo:
                        continue

                    try:
                        nuevo_stock = int(row['NuevoStock'])
                    except (TypeError, ValueError):
                        continue

                    try:
                        material = Material.objects.get(codigo=codigo)
                    except Material.DoesNotExist:
                        no_encontrados.append(codigo)
                        continue

                    inv, _ = Inventario.objects.get_or_create(material=material)

                    if modo == 'ajuste':
                        diferencia = nuevo_stock - inv.stock_actual
                        inv.stock_actual = nuevo_stock
                        tipo_mov = 'ajuste'
                        detalle = f"Ajuste masivo desde archivo. Nuevo stock={nuevo_stock}"
                    else:  # entrada
                        diferencia = nuevo_stock
                        inv.stock_actual += nuevo_stock
                        tipo_mov = 'entrada'
                        detalle = f"Entrada masiva desde archivo. +{nuevo_stock}"

                    inv.save()

                    if diferencia != 0:
                        Movimiento.objects.create(
                            material=material,
                            usuario=request.user,
                            tipo=tipo_mov,
                            cantidad=abs(diferencia),
                            detalle=detalle,
                        )

                    actualizados += 1

            messages.success(
                request,
                f"Stock actualizado para {actualizados} materiales. No encontrados: {len(no_encontrados)}"
            )
            return redirect('inventario')
    else:
        form = CargaMasivaStockForm()

    return render(request, 'funcionalidad/inv_carga_masiva_stock.html', {'form': form})


@login_required
@verificar_rol(['SISTEMA']) 
@require_POST
def recalcular_stock_ml(request):
    print(f"DEBUG: Datos recibidos -> {request.POST}") # Ver en consola

    # Fórmula
    formula = request.POST.get("formula", "conservadora")
    usar_conservadora = (formula == "conservadora")

    # Estación
    estacion_opcion = request.POST.get("estacion_manual", "")
    print(f"DEBUG: Opción elegida -> '{estacion_opcion}'")

    if estacion_opcion == "sin_estacion":
        usar_estacion = False
        estacion_manual = None
        print("DEBUG: Sin estación (usar todo el año)")
        
    elif estacion_opcion in ["Verano", "Otoño", "Invierno", "Primavera"]:
        usar_estacion = True
        estacion_manual = estacion_opcion
        print(f"DEBUG: Forzando estación MANUAL -> {estacion_manual}")
        
    else:
        # Caso "Detectar automáticamente" (value="")
        usar_estacion = True
        estacion_manual = None 
        print("DEBUG: Automático (Usará fecha del servidor)")

    # Días e Historial
    try:
        dias_historial = int(request.POST.get("dias_historial", "180"))
        nivel_servicio = float(request.POST.get("nivel_servicio", "0.95"))
    except ValueError:
        dias_historial = 180
        nivel_servicio = 0.95

    # Ejecutar
    resultados = ejecutar_calculo_global(
        usar_formula_conservadora=usar_conservadora,
        usar_estacion=usar_estacion,
        estacion_manual=estacion_manual, # <--- IMPORTANTE
        dias_historial=dias_historial,
        nivel_servicio=nivel_servicio,
    )

    count = len(resultados) if resultados else 0
    
    # Mensaje de éxito
    msg_estacion = estacion_manual if estacion_manual else ("Automática" if usar_estacion else "Desactivada")
    
    messages.success(
        request,
        f"Stock recalculado: {count} mats | Estación: {msg_estacion} | Historial: {dias_historial}d"
    )
    
    return redirect("prediccion_stock")


@login_required
@verificar_rol('SISTEMA')
def prediccion_stock(request):
    total_materiales = Material.objects.count()
    
    # Obtener TODOS los resultados ordenados por fecha (el más reciente primero)
    todos_resultados = MLResult.objects.select_related('material', 'material__inventario').order_by('-fecha_calculo')
    
    # Filtrar en Python para quedarnos solo con el último de cada material
    # Usamos un dict para rastrear si ya procesamos ese material
    resultados_unicos = {}
    ultimo_parametro = None

    for res in todos_resultados:
        if res.material.codigo not in resultados_unicos:
            resultados_unicos[res.material.codigo] = res
            # Guardamos el primer resultado (el más reciente de todos) para sacar info de parámetros
            if ultimo_parametro is None:
                ultimo_parametro = res
    
    # Convertir a lista
    resultados = list(resultados_unicos.values())
    
    # Procesar para la tabla
    tabla_resultados = []
    en_riesgo = 0
    
    for res in resultados:
        try:
            inv = res.material.inventario
            stock_actual = inv.stock_actual
            stock_critico = res.stock_min_calculado
            diferencia = stock_actual - stock_critico
            
            estado = 'OK'
            clase_css = 'success'
            
            if stock_actual <= 0:
                estado = 'QUIEBRE'
                clase_css = 'dark'
                en_riesgo += 1
            elif stock_actual < stock_critico:
                estado = 'CRÍTICO'
                clase_css = 'danger'
                en_riesgo += 1
            elif stock_actual < (stock_critico * 1.2):
                estado = 'ALERTA'
                clase_css = 'warning'
            
            tabla_resultados.append({
                'codigo': res.material.codigo,
                'descripcion': res.material.descripcion,
                'demanda_promedio': res.demanda_promedio,
                'desviacion': res.desviacion,
                'stock_critico': stock_critico,
                'stock_actual': stock_actual,
                'estado': estado,
                'clase_css': clase_css,
                'diferencia': diferencia,
                'stock_seguridad': getattr(res, 'stock_seguridad', 0),
                'coeficiente_variacion': getattr(res, 'coeficiente_variacion', 0),
                'metodo': getattr(res, 'metodo_utilizado', ''),
                'demanda_leadtime': res.stock_min_calculado - getattr(res, 'stock_seguridad', 0),
            })
        except Inventario.DoesNotExist:
            continue

    
    info_calculo = {
        'fecha': ultimo_parametro.fecha_calculo if ultimo_parametro else None,
        'modelo': ultimo_parametro.version_modelo if ultimo_parametro else 'N/A',
        
    }

    tabla_resultados.sort(key=lambda x: x['diferencia'])

    context = {
        'total_materiales': total_materiales,
        'materiales_calculados': len(tabla_resultados),
        'total_en_riesgo': en_riesgo,
        'tabla_resultados': tabla_resultados,
        'info_calculo': info_calculo
    }
    
    return render(request, 'funcionalidad/prediccion_stock.html', context)



# ============================================= SOLICITUDES DE MATERIALES ====================================
@login_required
@verificar_rol('TECNICO')
def crear_solicitud(request):
    
    if request.method == 'POST':
        form = SolicitudForm(request.POST)
        formset = DetalleSolicitudFormSet(request.POST)
        
        if form.is_valid() and formset.is_valid():
            
            detalles_validos = [
                 f for f in formset
                if f.cleaned_data
                and not f.cleaned_data.get('DELETE', False)
                and f.cleaned_data.get('material')
                and f.cleaned_data.get('cantidad')
            ]
            for f in formset:
                if f.cleaned_data and not f.cleaned_data.get('DELETE', False):
                    material = f.cleaned_data.get('material')
                    cantidad = f.cleaned_data.get('cantidad')
                    
                    # Solo agregar si tiene AMBOS campos llenos
                    if material and cantidad:
                        detalles_validos.append(f)
            
            if len(detalles_validos) == 0:
                messages.error(request, 'Debes solicitar al menos 1 material.')
                return render(request, 'funcionalidad/solmat_crear_solicitud.html', {
                    'form': form,
                    'formset': formset,
                })
            
            # Validar máximo 10 materiales
            if len(detalles_validos) > 10:
                messages.error(request, 'No puedes solicitar más de 10 materiales diferentes por solicitud.')
                return render(request, 'funcionalidad/solmat_crear_solicitud.html', {
                    'form': form,
                    'formset': formset,
                })
            
            # Validar cada cantidad máximo 10
            for detalle in detalles_validos:
                cantidad = detalle.cleaned_data.get('cantidad', 0)
                if cantidad > 10:
                    messages.error(
                        request, 
                        f'No puedes solicitar más de 10 unidades del material "{detalle.cleaned_data["material"]}". Solicitaste: {cantidad}'
                    )
                    return render(request, 'funcionalidad/solmat_crear_solicitud.html', {
                        'form': form,
                        'formset': formset,
                    })
            
            try:
                with transaction.atomic():
                    solicitud = form.save(commit=False)
                    solicitud.solicitante = request.user
                    solicitud.save()
                    
                    # GUARDAR: Solo los detalles válidos (con datos)
                    for detalle_form in detalles_validos:
                        detalle = detalle_form.save(commit=False)
                        detalle.solicitud = solicitud
                        detalle.save()
                    
                    cantidad_total = sum(d.cleaned_data.get('cantidad', 0) for d in detalles_validos)
                    
                    # Notificar a BODEGA
                    usuarios_bodega = Usuario.objects.filter(rol='BODEGA')
                    for bodeguero in usuarios_bodega:
                        Notificacion.objects.create(
                            usuario=bodeguero,
                            tipo='solicitud_pendiente',
                            mensaje=f'Nueva solicitud #{solicitud.id} de {request.user.get_full_name()} - {len(detalles_validos)} materiales ({cantidad_total} items)',
                            url=f'/solicitud/{solicitud.id}/'
                        )
                    
                    messages.success(
                        request, 
                        f'Solicitud #{solicitud.id} creada con {len(detalles_validos)} materiales ({cantidad_total} ítems totales).'
                    )
                    return redirect('mis_solicitudes')
            except Exception as e:
                messages.error(request, f'Error al crear la solicitud: {str(e)}')
        else:
            messages.error(request, 'Por favor corrige los errores del formulario.')
    else:
        form = SolicitudForm()
        formset = DetalleSolicitudFormSet()
    
    context = {
        'form': form,
        'formset': formset,
    }
    
    return render(request, 'funcionalidad/solmat_crear_solicitud.html', context)



@login_required
def mis_solicitudes(request):
    
    solicitudes = Solicitud.objects.filter(
        solicitante=request.user
    ).prefetch_related('detalles__material').order_by('-fecha_solicitud')
    
    # Estadísticas
    stats = {
        'total': solicitudes.count(),
        'pendientes': solicitudes.filter(estado='pendiente').count(),
        'aprobadas': solicitudes.filter(estado='aprobada').count(),
        'rechazadas': solicitudes.filter(estado='rechazada').count(),
        'despachadas': solicitudes.filter(estado='despachada').count(),
    }
    
    paginator = Paginator(solicitudes, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'solicitudes': page_obj, 
        'page_obj': page_obj,     
        'stats': stats,
    }
    
    return render(request, 'funcionalidad/solmat_mis_solicitudes.html', context)


@login_required
def detalle_solicitud(request, solicitud_id):
    
    solicitud = get_object_or_404(
        Solicitud.objects.prefetch_related('detalles__material'),
        id=solicitud_id
    )
    
    # Verificar permisos según rol
    if request.user.rol == 'TECNICO':
        if solicitud.solicitante != request.user:
            messages.error(request, 'No tienes permiso para ver esta solicitud.')
            return redirect('mis_solicitudes')
    elif request.user.rol in ['BODEGA', 'GERENCIA']:
        pass
    else:
        messages.error(request, 'No tienes permiso para ver solicitudes.')
        return redirect('dashboard')
    
    
    detalles_info = []
    tiene_stock_suficiente = True
    
    for detalle in solicitud.detalles.all():
        try:
            inventario = detalle.material.inventario
            stock_disponible = inventario.stock_actual
            nuevo_stock = stock_disponible - detalle.cantidad
            if nuevo_stock < 0:
                tiene_stock_suficiente = False
        except Inventario.DoesNotExist:
            stock_disponible = 0
            nuevo_stock = 0
            tiene_stock_suficiente = False
        
        detalles_info.append({
            'detalle': detalle,
            'stock_disponible': stock_disponible,
            'nuevo_stock': nuevo_stock
        })
    
    
    referer = request.META.get('HTTP_REFERER', '')
    
    if 'notificaciones' in referer:
        url_volver = reverse('mis_notificaciones')
        texto_volver = 'Volver a notificaciones'
    elif request.user.rol == 'TECNICO':
        url_volver = reverse('mis_solicitudes')
        texto_volver = 'Volver a mis solicitudes'
    elif request.user.rol in ['BODEGA', 'GERENCIA']:
        url_volver = reverse('gestionar_solicitudes')
        texto_volver = 'Volver a gestionar'
    else:
        url_volver = reverse('dashboard')
        texto_volver = 'Volver al inicio'
    
    context = {
        'solicitud': solicitud,
        'detalles_info': detalles_info,
        'tiene_stock_suficiente': tiene_stock_suficiente,
        'es_bodega_gerencia': request.user.rol in ['BODEGA', 'GERENCIA'],
        'es_bodega': request.user.rol == 'BODEGA',
        'url_volver': url_volver,  
        'texto_volver': texto_volver,  
    }
    
    return render(request, 'funcionalidad/solmat_detalle_solicitud.html', context)

@login_required
@verificar_rol(['BODEGA']) 
def gestionar_solicitudes(request):
    """Vista para que BODEGA gestione todas las solicitudes"""
    solicitudes_pendientes = Solicitud.objects.filter(
        estado='pendiente'
    ).prefetch_related('detalles__material').order_by('-fecha_solicitud')
    
    todas_solicitudes = Solicitud.objects.all().prefetch_related(
        'detalles__material'
    ).order_by('-fecha_solicitud')[:20]
    
    context = {
        'solicitudes_pendientes': solicitudes_pendientes,
        'todas_solicitudes': todas_solicitudes,
    }
    
    return render(request, 'funcionalidad/solmat_gestionar.html', context)

@login_required
def detalle_solicitud(request, solicitud_id):
    """Vista para ver el detalle completo de una solicitud"""
    solicitud = get_object_or_404(
        Solicitud.objects.prefetch_related('detalles__material'),
        id=solicitud_id
    )
    
    if request.user.rol == 'TECNICO':
        if solicitud.solicitante != request.user:
            messages.error(request, 'No tienes permiso para ver esta solicitud.')
            return redirect('mis_solicitudes')
    elif request.user.rol in ['BODEGA', 'GERENCIA']:
        pass
    else:
        messages.error(request, 'No tienes permiso para ver solicitudes.')
        return redirect('dashboard')
    
    detalles_info = []
    tiene_stock_suficiente = True
    
    for detalle in solicitud.detalles.all():
        try:
            inventario = detalle.material.inventario
            stock_disponible = inventario.stock_actual
            nuevo_stock = stock_disponible - detalle.cantidad
            if nuevo_stock < 0:
                tiene_stock_suficiente = False
        except Inventario.DoesNotExist:
            stock_disponible = 0
            nuevo_stock = 0
            tiene_stock_suficiente = False
        
        detalles_info.append({
            'detalle': detalle,
            'stock_disponible': stock_disponible,
            'nuevo_stock': nuevo_stock
        })
    
    context = {
        'solicitud': solicitud,
        'detalles_info': detalles_info,
        'tiene_stock_suficiente': tiene_stock_suficiente,
        'es_bodega_gerencia': request.user.rol in ['BODEGA', 'GERENCIA'],
        'es_bodega': request.user.rol == 'BODEGA'
    }
    
    
    return render(request, 'funcionalidad/solmat_detalle_solicitud.html', context)

        
@login_required
@verificar_rol('BODEGA')
def aprobar_solicitud(request, solicitud_id):
    
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)
    
    if solicitud.estado != 'pendiente':
        messages.error(request, 'Solo solicitudes pendientes.')
        return redirect('gestionar_solicitudes')
    
    if request.method == 'POST':
        with transaction.atomic():
            # VERIFICAR STOCK PRIMERO (separar verificación de descuento)
            for detalle in solicitud.detalles.all():
                inventario = detalle.material.inventario
                if inventario.stock_actual < detalle.cantidad:
                    messages.error(request, f'Stock insuficiente: {detalle.material.descripcion} (Disp: {inventario.stock_actual})')
                    return redirect('detalle_solicitud', solicitud_id=solicitud.id)
    
            # DESCONTAR STOCK (segunda pasada)
            for detalle in solicitud.detalles.all():
                inventario = detalle.material.inventario
                inventario.stock_actual -= detalle.cantidad
                inventario.save()
                # MOVIMIENTO
                Movimiento.objects.create(
                    material=detalle.material,
                    usuario=request.user,
                    solicitud=solicitud,
                    tipo='salida',
                    cantidad=detalle.cantidad,
                    detalle=f'Aprobación solicitud #{solicitud.id}'
                )
            
            # APROBAR FINAL
            solicitud.estado = 'aprobada'
            solicitud.respondido_por = request.user
            solicitud.fecha_respuesta = timezone.now()
            solicitud.save()
            
            # NOTIFICACIÓN
            Notificacion.objects.create(
                usuario=solicitud.solicitante,
                tipo='solicitud_aprobada',
                mensaje=f'Tu solicitud #{solicitud.id} ha sido APROBADA',
                url=f'/solicitud/{solicitud.id}/'
            )
            messages.success(request, f'Solicitud #{solicitud.id} aprobada y stock descontado.')
        return redirect('gestionar_solicitudes')
    
    return redirect('detalle_solicitud', solicitud_id=solicitud.id)



@login_required
@verificar_rol(['BODEGA'])
def rechazar_solicitud(request, solicitud_id):
   
    solicitud = get_object_or_404(Solicitud, id=solicitud_id)
    
    if solicitud.estado != 'pendiente':
        messages.error(request, 'Solo se pueden rechazar solicitudes pendientes.')
        return redirect('detalle_solicitud', solicitud_id=solicitud_id)
    
    if request.method == 'POST':
        observaciones = request.POST.get('observaciones', '').strip()
        
        # Validar que haya observaciones
        if not observaciones:
            messages.error(request, 'Debes ingresar una justificación para rechazar la solicitud.')
            return redirect('detalle_solicitud', solicitud_id=solicitud_id)
        
        if len(observaciones) < 10:
            messages.error(request, 'La justificación debe tener al menos 10 caracteres.')
            return redirect('detalle_solicitud', solicitud_id=solicitud_id)
        
        # Rechazar solicitud
        solicitud.estado = 'rechazada'
        solicitud.respondido_por = request.user
        solicitud.fecha_respuesta = timezone.now()
        solicitud.observaciones = observaciones
        solicitud.save()
        
        # CREAR NOTIFICACIÓN PARA EL TÉCNICO
        Notificacion.objects.create(
            usuario=solicitud.solicitante,
            tipo='solicitud_rechazada',
            mensaje=f'Tu solicitud #{solicitud.id} ha sido rechazada. Motivo: {observaciones[:100]}',
            url=f'/solicitud/{solicitud.id}/'
        )
        
        messages.warning(
            request, 
            f'Solicitud #{solicitud.id} rechazada. El técnico ha sido notificado.'
        )
        return redirect('gestionar_solicitudes')
    
    # GET: Mostrar formulario de rechazo
    context = {
        'solicitud': solicitud,
    }
    return render(request, 'funcionalidad/solmat_rechazar.html', context)


@login_required
@verificar_rol('TECNICO')
def cancelar_solicitud(request, solicitud_id):
    """Cancelar una solicitud propia (solo si está pendiente)"""
    solicitud = get_object_or_404(Solicitud, id=solicitud_id, solicitante=request.user)
    
    if solicitud.estado == 'pendiente' and request.method == 'POST':
        solicitud.delete()
        messages.info(request, 'Solicitud cancelada exitosamente.')
        return redirect('mis_solicitudes')
    
    messages.error(request, 'No se puede cancelar esta solicitud.')
    return redirect('detalle_solicitud', solicitud_id=solicitud_id)


@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def historial_solicitudes(request):
    from datetime import datetime
    
    # Filtros desde GET
    estado_filtro = request.GET.get('estado', '')
    solicitante_filtro = request.GET.get('solicitante', '')
    fecha_desde = request.GET.get('fecha_desde', '')
    fecha_hasta = request.GET.get('fecha_hasta', '')
    
    # Consulta base con agregación de cantidad total
    solicitudes = Solicitud.objects.select_related(
        'solicitante', 'respondido_por'
    ).prefetch_related('detalles__material').annotate(
        total_materiales=Sum('detalles__cantidad')  # ← AGREGAR ESTO
    ).order_by('-fecha_solicitud')
    
    # Aplicar filtros
    if estado_filtro:
        solicitudes = solicitudes.filter(estado=estado_filtro)
    
    if solicitante_filtro:
        solicitudes = solicitudes.filter(solicitante__username=solicitante_filtro)
    
    if fecha_desde:
        try:
            fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
            solicitudes = solicitudes.filter(fecha_solicitud__gte=fecha_desde_dt)
        except ValueError:
            pass
    
    if fecha_hasta:
        try:
            fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d')
            solicitudes = solicitudes.filter(fecha_solicitud__lte=fecha_hasta_dt)
        except ValueError:
            pass
    
    # Paginación
    paginator = Paginator(solicitudes, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas generales
    stats = {
        'total': Solicitud.objects.count(),
        'pendientes': Solicitud.objects.filter(estado='pendiente').count(),
        'aprobadas': Solicitud.objects.filter(estado='aprobada').count(),
        'rechazadas': Solicitud.objects.filter(estado='rechazada').count(),
        'canceladas': Solicitud.objects.filter(estado='cancelada').count(),
    }
    
    # Lista de técnicos para el filtro
    tecnicos = User.objects.filter(rol='TECNICO').order_by('first_name')
    
    context = {
        'solicitudes': page_obj,
        'page_obj': page_obj,
        'stats': stats,
        'tecnicos': tecnicos,
        'estado_filtro': estado_filtro,
        'solicitante_filtro': solicitante_filtro,
        'fecha_desde': fecha_desde,
        'fecha_hasta': fecha_hasta,
    }
    
    return render(request, 'funcionalidad/solmat_historial.html', context)


# ============================================= Registro de movimientos ====================================
@login_required
@verificar_rol(['BODEGA']) 
def registrar_entrada(request, material_id):
    
    material = get_object_or_404(Material, id=material_id)
    
    try:
        inventario = material.inventario
    except Inventario.DoesNotExist:
        messages.error(request, 'Este material no tiene inventario asociado.')
        return redirect('inventario')
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0))
        detalle = request.POST.get('detalle', '')
        
        if cantidad <= 0:
            messages.error(request, 'La cantidad debe ser mayor a 0.')
        else:
            try:
                with transaction.atomic():
                    stock_anterior = inventario.stock_actual
                    stock_nuevo = stock_anterior + cantidad
                    
                    inventario.stock_actual = stock_nuevo
                    inventario.save()
                    
                    try:
                        usuario_movimiento = Usuario.objects.get(
                            email=request.user.email
                        )
                    except Usuario.DoesNotExist:
                        messages.error(request, 'Tu usuario no está registrado en el sistema de personal.')
                        return redirect('detalle_material', id=material_id)
                    
                    Movimiento.objects.create(
                        material=material,
                        usuario=usuario_movimiento,
                        tipo='entrada',
                        cantidad=cantidad,
                        detalle=detalle or f'Entrada manual registrada por {request.user.username}'
                    )
                    
                    messages.success(
                        request,
                        f'Entrada registrada: +{cantidad} {material.unidad_medida}. Stock actual: {stock_nuevo}'
                    )
                    return redirect('detalle_material', id=material_id)
            except Exception as e:
                messages.error(request, f'Error al registrar entrada: {str(e)}')
    
    context = {
        'material': material,
        'inventario': inventario,
    }
    
    return render(request, 'funcionalidad/mov_registrar_entrada.html', context)

@login_required
@verificar_rol('BODEGA') 
def registrar_salida(request, material_id):
    """Registrar salida manual de material"""
    material = get_object_or_404(Material, id=material_id)
    
    try:
        inventario = material.inventario
    except Inventario.DoesNotExist:
        messages.error(request, 'Este material no tiene inventario asociado.')
        return redirect('inventario')
    
    if request.method == 'POST':
        cantidad = int(request.POST.get('cantidad', 0))
        detalle = request.POST.get('detalle', '')
        
        if cantidad <= 0:
            messages.error(request, 'La cantidad debe ser mayor a 0.')
        elif cantidad > inventario.stock_actual:
            messages.error(request, f'Stock insuficiente. Disponible: {inventario.stock_actual}')
        else:
            try:
                with transaction.atomic():
                    stock_anterior = inventario.stock_actual
                    stock_nuevo = stock_anterior - cantidad
                    
                    inventario.stock_actual = stock_nuevo
                    inventario.save()
                    
                    try:
                        usuario_movimiento = Usuario.objects.get(
                            email=request.user.email
                        )
                    except Usuario.DoesNotExist:
                        messages.error(request, 'Tu usuario no está registrado en el sistema de personal.')
                        return redirect('detalle_material', id=material_id)
                    
                    Movimiento.objects.create(
                        material=material,
                        usuario=usuario_movimiento,
                        tipo='salida',
                        cantidad=cantidad,
                        detalle=detalle or f'Salida manual registrada por {request.user.username}'
                    )
                    
                    messages.success(
                        request,
                        f'Salida registrada: -{cantidad} {material.unidad_medida}. Stock actual: {stock_nuevo}'
                    )
                    return redirect('detalle_material', id=material_id)
            except Exception as e:
                messages.error(request, f'Error al registrar salida: {str(e)}')
    
    context = {
        'material': material,
        'inventario': inventario,
    }
    
    return render(request, 'funcionalidad/mov_registrar_salida.html', context)

@login_required
@verificar_rol('BODEGA')
def ajustar_inventario(request, material_id):
    """Ajustar inventario (correcciones)"""
    material = get_object_or_404(Material, id=material_id)
    
    try:
        inventario = material.inventario
    except Inventario.DoesNotExist:
        messages.error(request, 'Este material no tiene inventario asociado.')
        return redirect('inventario')
    
    if request.method == 'POST':
        nuevo_stock = int(request.POST.get('nuevo_stock', 0))
        detalle = request.POST.get('detalle', '')
        
        if nuevo_stock < 0:
            messages.error(request, 'El stock no puede ser negativo.')
        else:
            try:
                with transaction.atomic():
                    stock_anterior = inventario.stock_actual
                    diferencia = nuevo_stock - stock_anterior
                    
                    inventario.stock_actual = nuevo_stock
                    inventario.save()
                    
                    try:
                        usuario_movimiento = Usuario.objects.get(
                            email=request.user.email
                        )
                    except Usuario.DoesNotExist:
                        messages.error(request, 'Tu usuario no está registrado en el sistema de personal.')
                        return redirect('detalle_material', id=material_id)
                    
                    Movimiento.objects.create(
                        material=material,
                        usuario=usuario_movimiento,
                        tipo='ajuste',
                        cantidad=abs(diferencia),
                        detalle=f'Ajuste de inventario ({diferencia:+d}): {detalle}'
                    )
                    
                    messages.success(
                        request,
                        f'Inventario ajustado. Stock anterior: {stock_anterior}, Stock nuevo: {nuevo_stock}'
                    )
                    return redirect('detalle_material', id=material_id)
            except Exception as e:
                messages.error(request, f'Error al ajustar inventario: {str(e)}')
    
    context = {
        'material': material,
        'inventario': inventario,
    }
    
    return render(request, 'funcionalidad/mov_ajustar_inventario.html', context)

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def historial_movimientos(request, material_id):
    """Ver historial de movimientos de un material"""
    material = get_object_or_404(Material, id=material_id)
    movimientos = Movimiento.objects.filter(
        material=material
    ).select_related('usuario').order_by('-fecha')
    
    context = {
        'material': material,
        'movimientos': movimientos,
    }
    
    return render(request, 'funcionalidad/mov_historial.html', context)

@login_required
@verificar_rol(['BODEGA', 'GERENCIA']) 
def historial_movimientos_global(request):
    """Historial completo de todos los movimientos del sistema"""
    movimientos = Movimiento.objects.select_related(
        'material', 'usuario'
    ).all().order_by('-fecha')[:100]
    
    context = {
        'movimientos': movimientos
    }
    
    return render(request, 'funcionalidad/mov_historial_global.html', context)


@login_required
def solicitud(request):
    """
    Vista simple de solicitud.
    Redirige a crear solicitud.
    """
    return redirect('crear_solicitud')

# ============================================= NOTIFICACIONES ====================================

@login_required
def mis_notificaciones(request):
    """
    Vista para listar todas las notificaciones del usuario.
    
    - TÉCNICO: Solo notificaciones de sus solicitudes
    - BODEGA/GERENCIA: Todas las notificaciones
    """
    usuario = request.user
    rol = usuario.rol
    
    # Filtrar notificaciones según el rol
    if rol == 'TECNICO':
        notificaciones = Notificacion.objects.filter(
            usuario=usuario,
            tipo__in=['solicitud_aprobada', 'solicitud_rechazada']
        ).order_by('-creada_en')
    elif rol in ['BODEGA', 'GERENCIA']:
        notificaciones = Notificacion.objects.filter(
            usuario=usuario
        ).order_by('-creada_en')
    else:
        notificaciones = Notificacion.objects.none()
    
    # Paginación
    paginator = Paginator(notificaciones, 20)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Estadísticas
    stats = {
        'total': notificaciones.count(),
        'no_leidas': notificaciones.filter(leida=False).count(),
        'leidas': notificaciones.filter(leida=True).count(),
    }
    
    context = {
        'notificaciones': page_obj,
        'page_obj': page_obj,
        'stats': stats,
    }
    
    return render(request, 'funcionalidad/notificaciones.html', context)

@login_required
def marcar_leida(request, id):
    
    notificacion = get_object_or_404(Notificacion, id=id, usuario=request.user)
    notificacion.leida = True
    notificacion.save()
    messages.success(request, 'Notificación marcada como leída.')
    return redirect('mis_notificaciones')

@login_required
def marcar_todas_leidas(request):
    
    if request.method == 'POST':
        count = Notificacion.objects.filter(
            usuario=request.user,
            leida=False
        ).update(leida=True)
        
        messages.success(request, f'✓ {count} notificaciones marcadas como leídas.')
    
    return redirect('mis_notificaciones')

@login_required
def leer_notificacion(request, id):
    
    notificacion = get_object_or_404(Notificacion, id=id, usuario=request.user)
    notificacion.leida = True
    notificacion.save()
    
    # Redirigir a la URL de la notificación si existe
    if notificacion.url:
        return redirect(notificacion.url)
    return redirect('mis_notificaciones')

@login_required
def eliminar_notificacion(request, id):
    
    notificacion = get_object_or_404(Notificacion, id=id, usuario=request.user)
    
    if request.method == 'POST':
        notificacion.delete()
        messages.success(request, '✓ Notificación eliminada.')
    
    return redirect('mis_notificaciones')

@login_required
def obtener_notificaciones_json(request):
    """
    API endpoint para obtener notificaciones no leídas en formato JSON.
    Usado por el dropdown del navbar.
    
    - TÉCNICO: Solo notificaciones de sus solicitudes
    - BODEGA/GERENCIA: Todas las notificaciones del sistema
    """
    usuario = request.user
    rol = usuario.rol
    
    # Filtrar notificaciones según el rol
    if rol == 'TECNICO':
        # TÉCNICO: Solo notificaciones relacionadas a SUS solicitudes
        notificaciones = Notificacion.objects.filter(
            usuario=usuario,
            leida=False,
            tipo__in=['solicitud_aprobada', 'solicitud_rechazada']
        ).order_by('-creada_en')[:10]
    elif rol in ['BODEGA', 'GERENCIA']:
        # BODEGA y GERENCIA: Todas las notificaciones
        notificaciones = Notificacion.objects.filter(
            usuario=usuario,
            leida=False
        ).order_by('-creada_en')[:10]
    else:
        # Otro rol: Sin notificaciones
        notificaciones = Notificacion.objects.none()
    
    # Contar total de no leídas
    if rol == 'TECNICO':
        count_total = Notificacion.objects.filter(
            usuario=usuario,
            leida=False,
            tipo__in=['solicitud_aprobada', 'solicitud_rechazada']
        ).count()
    elif rol in ['BODEGA', 'GERENCIA']:
        count_total = Notificacion.objects.filter(
            usuario=usuario,
            leida=False
        ).count()
    else:
        count_total = 0
    
    data = {
        'count': count_total,
        'notificaciones': [
            {
                'id': n.id,
                'tipo': n.tipo,
                'mensaje': n.mensaje,
                'url': n.url or '',
                'fecha': n.creada_en.strftime('%d/%m/%Y %H:%M'),
                'icono': get_icono_notificacion(n.tipo)
            }
            for n in notificaciones
        ]
    }
    
    return JsonResponse(data)

def get_icono_notificacion(tipo):
    """Retorna el ícono FontAwesome según el tipo de notificación"""
    iconos = {
        'stock_critico': 'fa-exclamation-triangle',
        'solicitud_pendiente': 'fa-clipboard-list',
        'material_nuevo': 'fa-box',
        'aprobacion': 'fa-check-circle',
        'solicitud_aprobada': 'fa-check-circle',      # Para técnicos
        'solicitud_rechazada': 'fa-times-circle',     # Para técnicos
    }
    return iconos.get(tipo, 'fa-bell')


# ============================================= DASHBOARD ====================================
@login_required
def dashboard(request):
    """Dashboard principal con estadísticas adaptadas según el rol del usuario."""
    usuario = request.user
    rol = usuario.rol
    
    # KPIs BASE para todos los roles
    total_materiales = Material.objects.count()
    solicitudes_totales = Solicitud.objects.count()
    solicitudes_pendientes = Solicitud.objects.filter(estado='pendiente').count()
    
    # Materiales críticos
    materiales_criticos_lista = Inventario.objects.filter(
        stock_actual__lte=F('stock_seguridad')
    ).select_related('material')[:5]
    materiales_criticos = Inventario.objects.filter(
        stock_actual__lte=F('stock_seguridad')
    ).count()
    
    # Context base
    context = {
        'usuario': usuario,
        'rol': rol,
        'total_materiales': total_materiales,
        'solicitudes_totales': solicitudes_totales,
        'solicitudes_pendientes': solicitudes_pendientes,
        'materiales_criticos': materiales_criticos,
        'materiales_criticos_lista': materiales_criticos_lista,
    }
    
    # ========== TÉCNICO ==========
    if rol == 'TECNICO':
        mis_solicitudes = Solicitud.objects.filter(solicitante=usuario)
        
        # Stats para gráfico de solicitudes
        solicitudes_data = [
            mis_solicitudes.filter(estado='pendiente').count(),
            mis_solicitudes.filter(estado='aprobada').count(),
            mis_solicitudes.filter(estado='rechazada').count(),
        ]
        
        context.update({
            'mis_solicitudes_count': mis_solicitudes.count(),
            'mis_pendientes': mis_solicitudes.filter(estado='pendiente').count(),
            'mis_aprobadas': mis_solicitudes.filter(estado='aprobada').count(),
            'solicitudes_recientes': mis_solicitudes.order_by('-fecha_solicitud')[:5],
            
            # Top materiales del técnico
            'top_materiales': DetalleSolicitud.objects.filter(
                solicitud__solicitante=usuario
            ).values(
                'material__codigo', 'material__descripcion'
            ).annotate(
                total_solicitado=Sum('cantidad')
            ).order_by('-total_solicitado')[:5],
            
            # Stats para gráfico - JSON serializado
            'solicitudes_labels': json.dumps(['Pendientes', 'Aprobadas', 'Rechazadas']),
            'solicitudes_data': json.dumps(solicitudes_data),
        })
    
    # ========== BODEGA ==========
    elif rol == 'BODEGA':
        # Stock total
        stock_total = Inventario.objects.aggregate(
            total=Sum('stock_actual')
        )['total'] or 0
        
        # Stats para gráfico de solicitudes
        solicitudes_data = [
            Solicitud.objects.filter(estado='pendiente').count(),
            Solicitud.objects.filter(estado='aprobada').count(),
            Solicitud.objects.filter(estado='rechazada').count(),
        ]
        
        # ========== PROCESAR MOVIMIENTOS DE LOS ÚLTIMOS 7 DÍAS ==========
        movimientos_7dias = Movimiento.objects.filter(
            fecha__gte=timezone.now() - timedelta(days=7)
        ).values('fecha__date', 'tipo').annotate(
            total=Sum('cantidad')
        ).order_by('fecha__date')
        
        # Agrupar por fecha y tipo
        mov_dict = defaultdict(lambda: {'entrada': 0, 'salida': 0})
        for mov in movimientos_7dias:
            fecha = mov['fecha__date'].strftime('%d/%m')
            mov_dict[fecha][mov['tipo']] = mov['total']
        
        # Obtener últimos 7 días
        fechas = []
        entradas = []
        salidas = []
        for i in range(6, -1, -1):
            fecha = (timezone.now() - timedelta(days=i)).date()
            fecha_str = fecha.strftime('%d/%m')
            fechas.append(fecha_str)
            entradas.append(mov_dict[fecha_str]['entrada'])
            salidas.append(mov_dict[fecha_str]['salida'])
        
        context.update({
            'stock_total': stock_total,
            'solicitudes_recientes': Solicitud.objects.select_related('solicitante').order_by('-fecha_solicitud')[:5],
            
            # Top materiales del sistema
            'top_materiales': DetalleSolicitud.objects.values(
                'material__codigo', 'material__descripcion'
            ).annotate(
                total_solicitado=Sum('cantidad')
            ).order_by('-total_solicitado')[:5],
            
            # Actividad reciente
            'actividad_reciente': Movimiento.objects.select_related(
                'material', 'usuario'
            ).order_by('-fecha')[:10],
            
            # Stats para gráficos - JSON serializado
            'solicitudes_labels': json.dumps(['Pendientes', 'Aprobadas', 'Rechazadas']),
            'solicitudes_data': json.dumps(solicitudes_data),
            
            # Datos para gráfico de movimientos
            'movimientos_fechas': json.dumps(fechas),
            'movimientos_entradas': json.dumps(entradas),
            'movimientos_salidas': json.dumps(salidas),
        })
    
    # ========== GERENCIA ==========
    elif rol == 'GERENCIA':
        total_usuarios = User.objects.filter(is_active=True).count()
        stock_total = Inventario.objects.aggregate(
            total=Sum('stock_actual')
        )['total'] or 0
        
        # Stats para gráfico de solicitudes
        solicitudes_data = [
            Solicitud.objects.filter(estado='pendiente').count(),
            Solicitud.objects.filter(estado='aprobada').count(),
            Solicitud.objects.filter(estado='rechazada').count(),
        ]
        
        # ========== PROCESAR MOVIMIENTOS DE LOS ÚLTIMOS 7 DÍAS ==========
        movimientos_7dias = Movimiento.objects.filter(
            fecha__gte=timezone.now() - timedelta(days=7)
        ).values('fecha__date', 'tipo').annotate(
            total=Sum('cantidad')
        ).order_by('fecha__date')
        
        # Agrupar por fecha y tipo
        mov_dict = defaultdict(lambda: {'entrada': 0, 'salida': 0})
        for mov in movimientos_7dias:
            fecha = mov['fecha__date'].strftime('%d/%m')
            mov_dict[fecha][mov['tipo']] = mov['total']
        
        # Obtener últimos 7 días
        fechas = []
        entradas = []
        salidas = []
        for i in range(6, -1, -1):
            fecha = (timezone.now() - timedelta(days=i)).date()
            fecha_str = fecha.strftime('%d/%m')
            fechas.append(fecha_str)
            entradas.append(mov_dict[fecha_str]['entrada'])
            salidas.append(mov_dict[fecha_str]['salida'])
        
        context.update({
            'total_usuarios': total_usuarios,
            'stock_total': stock_total,
            'solicitudes_recientes': Solicitud.objects.select_related('solicitante').order_by('-fecha_solicitud')[:5],
            
            # Top materiales del sistema
            'top_materiales': DetalleSolicitud.objects.values(
                'material__codigo', 'material__descripcion'
            ).annotate(
                total_solicitado=Sum('cantidad')
            ).order_by('-total_solicitado')[:5],
            
            # Actividad reciente
            'actividad_reciente': Movimiento.objects.select_related(
                'material', 'usuario'
            ).order_by('-fecha')[:10],
            
            # Stats para gráficos - JSON serializado
            'solicitudes_labels': json.dumps(['Pendientes', 'Aprobadas', 'Rechazadas']),
            'solicitudes_data': json.dumps(solicitudes_data),
            
            # Datos para gráfico de movimientos
            'movimientos_fechas': json.dumps(fechas),
            'movimientos_entradas': json.dumps(entradas),
            'movimientos_salidas': json.dumps(salidas),
        })
    
    # ========== MATERIALES POR CATEGORÍA (para todos) ==========
    materiales_por_categoria = Material.objects.values('categoria').annotate(
        total=Count('id')
    ).order_by('-total')
    
    categorias_labels = [dict(Material.CATEGORIA_CHOICES).get(item['categoria'], item['categoria']) 
                         for item in materiales_por_categoria]
    categorias_data = [item['total'] for item in materiales_por_categoria]
    
    context.update({
        'categorias_labels': json.dumps(categorias_labels),
        'categorias_data': json.dumps(categorias_data),
    })
    
    return render(request, 'general/dashboard.html', context)



# ============================================= EXPORTACIONES A EXCEL ====================================

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])  # Ambos pueden exportar
def exportar_inventario_excel(request):
    """Exportar inventario completo a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"
    
    # Headers
    headers = ['Código', 'Descripción', 'Categoría', 'Stock Actual', 'Stock Seguridad', 'Estado', 'Ubicación']
    ws.append(headers)
    
    # Estilo para headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    inventario = Inventario.objects.select_related('material').all()
    for inv in inventario:
        if inv.stock_actual <= inv.stock_seguridad:
            estado = 'CRÍTICO'
        else:
            estado = 'NORMAL'
        
        ws.append([
            inv.material.codigo,
            inv.material.descripcion,
            inv.material.get_categoria_display(),
            inv.stock_actual,
            inv.stock_seguridad,
            estado,
            inv.material.ubicacion or 'No especificada'
        ])
    
    # Ajustar anchos
    for col in range(1, 8):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=inventario_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def exportar_solicitudes_excel(request):
    """Exportar solicitudes a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Solicitudes"
    
    # Headers
    headers = ['ID', 'Fecha', 'Solicitante', 'Estado', 'Total Items', 'Motivo']
    ws.append(headers)
    
    # Estilo para headers
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Datos
    solicitudes = Solicitud.objects.select_related('solicitante').prefetch_related('detalles').order_by('-fecha_solicitud')[:100]
    for sol in solicitudes:
        ws.append([
            sol.id,
            sol.fecha_solicitud.strftime('%d/%m/%Y %H:%M'),
            sol.solicitante.username,
            sol.get_estado_display(),
            sol.total_items(),
            sol.motivo[:50]
        ])
    
    # Ajustar anchos
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=solicitudes_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response

@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def exportar_movimientos_excel(request, material_id):
    """Exportar movimientos de UN material a Excel"""
    material = get_object_or_404(Material, id=material_id)

    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Movimientos"

    # Encabezados
    headers = [
        'ID', 'Fecha', 'Material', 'Código', 'Tipo',
        'Cantidad', 'Usuario', 'Detalle', 'Solicitud'
    ]

    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # SOLO movimientos de este material
    movimientos = Movimiento.objects.select_related(
        'material', 'usuario', 'solicitud'
    ).filter(material=material).order_by('-fecha')

    for row_num, mov in enumerate(movimientos, 2):
        if mov.usuario:
            usuario_nombre = mov.usuario.get_full_name() or mov.usuario.username
        else:
            usuario_nombre = 'Sistema'

        solicitud_id = f"#{mov.solicitud.id}" if mov.solicitud else "-"
        tipo_display = mov.get_tipo_display() if hasattr(mov, 'get_tipo_display') else mov.tipo

        ws.append([
            mov.id,
            mov.fecha.strftime('%d/%m/%Y %H:%M') if mov.fecha else '',
            mov.material.descripcion,
            mov.material.codigo,
            tipo_display.upper(),
            mov.cantidad,
            usuario_nombre,
            mov.detalle or '-',
            solicitud_id
        ])

    column_widths = {
        'A': 8,   # ID
        'B': 18,  # Fecha
        'C': 35,  # Material
        'D': 12,  # Código
        'E': 12,  # Tipo
        'F': 10,  # Cantidad
        'G': 20,  # Usuario
        'H': 40,  # Detalle
        'I': 12   # Solicitud
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            if cell.column == 5:  # Columna Tipo
                if cell.value == 'ENTRADA':
                    cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    cell.font = Font(color="006100", bold=True)
                elif cell.value == 'SALIDA':
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                    cell.font = Font(color="9C0006", bold=True)
                elif cell.value == 'AJUSTE':
                    cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    cell.font = Font(color="9C5700", bold=True)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f'movimientos_{material.codigo}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response



@login_required
@verificar_rol(['BODEGA', 'GERENCIA'])
def exportar_reporte_completo_excel(request):
    """Reporte completo con múltiples hojas: Inventario, Solicitudes, Movimientos"""
    wb = Workbook()
    wb.remove(wb.active)
    
    # ========== HOJA 1: INVENTARIO ==========
    ws_inv = wb.create_sheet("Inventario")
    headers_inv = ['Código', 'Descripción', 'Categoría', 'Stock Actual', 'Stock Seguridad', 'Estado']
    ws_inv.append(headers_inv)
    
    # Estilo headers
    for cell in ws_inv[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    inventario = Inventario.objects.select_related('material').all()
    for inv in inventario:
        estado = 'CRÍTICO' if inv.stock_actual <= inv.stock_seguridad else 'NORMAL'
        ws_inv.append([
            inv.material.codigo,
            inv.material.descripcion,
            inv.material.get_categoria_display(),
            inv.stock_actual,
            inv.stock_seguridad,
            estado
        ])
    
    # ========== HOJA 2: SOLICITUDES ==========
    ws_sol = wb.create_sheet("Solicitudes")
    headers_sol = ['ID', 'Fecha', 'Solicitante', 'Estado', 'Items']
    ws_sol.append(headers_sol)
    
    for cell in ws_sol[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    
    solicitudes = Solicitud.objects.select_related('solicitante').prefetch_related('detalles').order_by('-fecha_solicitud')[:100]
    for sol in solicitudes:
        ws_sol.append([
            sol.id,
            sol.fecha_solicitud.strftime('%d/%m/%Y'),
            sol.solicitante.username,
            sol.get_estado_display(),
            sol.total_items()
        ])
    
    # Preparar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=reporte_completo_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(response)
    return response


 # ==================================== GESTION DE LOCALES ==============================================

@login_required
@verificar_rol('SISTEMA')
def gestion_locales(request):
    
    locales = Local.objects.all().order_by('codigo', 'nombre')
    
    # Búsqueda
    query = request.GET.get('q', '').strip()
    if query:
        locales = locales.filter(
            Q(nombre__icontains=query) | 
            Q(codigo__icontains=query) |
            Q(comuna__icontains=query)
        )
    
    # Paginación
    paginator = Paginator(locales, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'locales': page_obj,
        'page_obj': page_obj,
        'query': query,
        'total_locales': locales.count()
    }
    return render(request, 'funcionalidad/local_gestion.html', context)

@login_required
@verificar_rol('SISTEMA')
def local_crear(request):
    
    if request.method == 'POST':
        form = LocalForm(request.POST)
        if form.is_valid():
            local = form.save()
            messages.success(
                request, 
                f'✓ Local {local.codigo} - {local.nombre} creado exitosamente.'
            )
            return redirect('gestion_locales')
    else:
        form = LocalForm()
    
    context = {
        'form': form,
        'titulo': 'Agregar Local',
        'accion': 'Crear'
    }
    return render(request, 'funcionalidad/local_form.html', context)

@login_required
@verificar_rol('SISTEMA')
def local_editar(request, local_id):
    
    local = get_object_or_404(Local, pk=local_id)
    
    if request.method == 'POST':
        form = LocalForm(request.POST, instance=local)
        if form.is_valid():
            form.save()
            messages.success(
                request, 
                f'✓ Local {local.codigo} - {local.nombre} actualizado exitosamente.'
            )
            return redirect('gestion_locales')
    else:
        form = LocalForm(instance=local)
    
    context = {
        'form': form,
        'local': local,
        'titulo': 'Editar Local',
        'accion': 'Actualizar'
    }
    return render(request, 'funcionalidad/local_form.html', context)

@login_required
@verificar_rol('SISTEMA')
def local_eliminar(request, local_id):
    
    local = get_object_or_404(Local, pk=local_id)
    
    if request.method == 'POST':
        # Verificar si tiene solicitudes asociadas
        if local.solicitudes.exists():
            messages.error(
                request,
                f'No se puede eliminar el local {local.codigo} - {local.nombre} '
                f'porque tiene {local.solicitudes.count()} solicitudes asociadas.'
            )
            return redirect('gestion_locales')
        
        nombre = str(local)
        local.delete()
        messages.success(request, f'✓ Local {nombre} eliminado exitosamente.')
        return redirect('gestion_locales')
    
    context = {
        'local': local,
        'solicitudes_count': local.solicitudes.count()
    }
    return render(request, 'funcionalidad/local_confirm_delete.html', context)


# ==========================================  GESTIÓN DE USUARIOS ==========================================

@login_required
@verificar_rol('SISTEMA')
def gestion_usuarios(request):
   
    usuarios = Usuario.objects.all().order_by('-date_joined')
    
    # Búsqueda
    query = request.GET.get('q', '').strip()
    if query:
        usuarios = usuarios.filter(
            Q(username__icontains=query) |
            Q(email__icontains=query) |
            Q(first_name__icontains=query) |
            Q(last_name__icontains=query)
        )
    
    # Filtro por rol
    rol_filtro = request.GET.get('rol', '')
    if rol_filtro:
        usuarios = usuarios.filter(rol=rol_filtro)
    
    # Filtro por estado
    estado_filtro = request.GET.get('estado', '')
    if estado_filtro == 'activo':
        usuarios = usuarios.filter(is_active=True)
    elif estado_filtro == 'inactivo':
        usuarios = usuarios.filter(is_active=False)
    
    # Estadísticas
    stats = {
        'total': Usuario.objects.count(),
        'activos': Usuario.objects.filter(is_active=True).count(),
        'inactivos': Usuario.objects.filter(is_active=False).count(),
        'tecnicos': Usuario.objects.filter(rol='TECNICO').count(),
        'bodega': Usuario.objects.filter(rol='BODEGA').count(),
        'gerencia': Usuario.objects.filter(rol='GERENCIA').count(),
        'gerencia': Usuario.objects.filter(rol='SISTEMA').count(),
    }
    
    # Paginación
    paginator = Paginator(usuarios, 15)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'usuarios': page_obj,
        'page_obj': page_obj,
        'query': query,
        'rol_filtro': rol_filtro,
        'estado_filtro': estado_filtro,
        'stats': stats,
        'roles': Usuario.ROL_CHOICES,
    }
    
    return render(request, 'funcionalidad/usuarios_gestion.html', context)


@login_required
@verificar_rol('SISTEMA')
def usuario_crear(request):
    
    if request.method == 'POST':
        form = UsuarioForm(request.POST, is_new=True)
        
        if form.is_valid():
            usuario = form.save()
            messages.success(
                request, 
                f'Usuario "{usuario.username}" creado exitosamente con rol {usuario.get_rol_display()}.'
            )
            return redirect('gestion_usuarios')
    else:
        form = UsuarioForm(is_new=True)
    
    context = {
        'form': form,
        'titulo': 'Crear Usuario',
        'accion': 'Crear'
    }
    
    return render(request, 'funcionalidad/usuarios_form.html', context)


@login_required
@verificar_rol('SISTEMA')
def usuario_editar(request, usuario_id):
    
    usuario = get_object_or_404(Usuario, pk=usuario_id)
    
    # Prevenir que se edite a sí mismo (opcional)
    if usuario == request.user:
        messages.warning(request, 'No puedes editar tu propio usuario. Usa "Cambiar Contraseña" en tu perfil.')
        return redirect('gestion_usuarios')
    
    if request.method == 'POST':
        form = UsuarioForm(request.POST, instance=usuario, is_new=False)
        
        if form.is_valid():
            usuario = form.save()
            messages.success(
                request, 
                f'Usuario "{usuario.username}" actualizado exitosamente.'
            )
            return redirect('gestion_usuarios')
    else:
        form = UsuarioForm(instance=usuario, is_new=False)
    
    context = {
        'form': form,
        'usuario': usuario,
        'titulo': f'Editar Usuario: {usuario.username}',
        'accion': 'Actualizar'
    }
    
    return render(request, 'funcionalidad/usuarios_form.html', context)


@login_required
@verificar_rol('SISTEMA')
def usuario_eliminar(request, usuario_id):
    
    usuario = get_object_or_404(Usuario, pk=usuario_id)
    
    # Prevenir que se elimine a sí mismo
    if usuario == request.user:
        messages.error(request, 'No puedes eliminar tu propio usuario.')
        return redirect('gestion_usuarios')
    
    if request.method == 'POST':
        # Verificar si tiene solicitudes asociadas
        solicitudes_count = Solicitud.objects.filter(solicitante=usuario).count()
        
        if solicitudes_count > 0:
            # No eliminar, solo desactivar
            usuario.is_active = False
            usuario.save()
            messages.warning(
                request, 
                f'Usuario "{usuario.username}" desactivado (tiene {solicitudes_count} solicitudes asociadas).'
            )
        else:
            # Eliminar permanentemente
            nombre = f"{usuario.username}"
            usuario.delete()
            messages.success(request, f'Usuario "{nombre}" eliminado exitosamente.')
        
        return redirect('gestion_usuarios')
    
    # Contar solicitudes
    solicitudes_count = Solicitud.objects.filter(solicitante=usuario).count()
    
    context = {
        'usuario': usuario,
        'solicitudes_count': solicitudes_count,
    }
    
    return render(request, 'funcionalidad/usuarios_confirm_delete.html', context)


@login_required
@verificar_rol('SISTEMA')
def usuario_toggle_estado(request, usuario_id):
    """Activar/desactivar usuario rápidamente"""
    if request.method == 'POST':
        usuario = get_object_or_404(Usuario, pk=usuario_id)
        
        if usuario == request.user:
            messages.error(request, 'No puedes cambiar tu propio estado.')
            return redirect('gestion_usuarios')
        
        usuario.is_active = not usuario.is_active
        usuario.save()
        
        estado = "activado" if usuario.is_active else "desactivado"
        messages.success(request, f'Usuario "{usuario.username}" {estado}.')
    
    return redirect('gestion_usuarios')





# ---------------- PANEL SISTEMA / SUPERUSUARIO ----------------
@login_required
def sistema_home(request):
    """
    Pantalla de inicio para el rol SISTEMA / superusuario.
    Si no es sistema ni superuser, se redirige al dashboard normal.
    """
    usuario = request.user

    if not (usuario.is_superuser or usuario.rol == 'SISTEMA'):
        return redirect('dashboard')

    # Stats básicos que le interesan al admin del sistema
    total_usuarios = Usuario.objects.count()
    activos = Usuario.objects.filter(is_active=True).count()
    inactivos = Usuario.objects.filter(is_active=False).count()
    total_locales = Local.objects.count()
    total_materiales = Material.objects.count()
    solicitudes_totales = Solicitud.objects.count()

    context = {
        'total_usuarios': total_usuarios,
        'activos': activos,
        'inactivos': inactivos,
        'total_locales': total_locales,
        'total_materiales': total_materiales,
        'solicitudes_totales': solicitudes_totales,
    }
    return render(request, 'rol/sistema.html', context)